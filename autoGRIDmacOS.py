from openpyxl import load_workbook # for opening xlsx files like the grids
from openpyxl.styles import PatternFill # to fill columns
from openpyxl.styles.borders import Border, Side # to create borders an new columns
from openpyxl.styles import Font # to write the bold number at the top of columns
import csv # for opening csv  files like the attendance reports
import string # for removing symbols from names

# create global variable to store the first open line in case we have to add any names to the grid
lastRow = int(input('\nNumber of the first blank row: '))
# create global variable to store the line which has the totals at the bottom so we don't overwrite it
totalsRow = int(input('Number of the last row with the totals on it: '))
# create global variable to know what row names not listed is on
namesNotListed = int(input("Number of the row that says 'names not listed': "))
# create global variable to store the names of any mentors that may have observed the SI leader so we don't put their names in the grid
mentors = {}
# ask how many times they were observed
numObservations = int(input("\nNumber of times you were observed this month: "))
for i in range(0, numObservations):
    if i > 0:
        print('')
    temp1 = str(input("\tFirst name of mentor #{}: ".format(i+1)))
    temp2 = str(input("\tLast name of mentor #{}: ".format(i+1)))
    # add the first name of the mentor as a key in the dictionary and the last name as the value
    mentors[temp1] = temp2
# global list to store any names that could not be placed in grid
unusedNames = []
# global dictionary to store nicknames
synonym = {}
# global dictionary to store non synonyms (people who have similar first names with the same last name, but are not the same person)
notSynonym = {}

# create dictionary of letters and numbers so we can find the proper letter for a certain column number
numToLetter = {} # numbers are keys and letters are values
# create string of uppercase letters
upperCaseString = string.ascii_uppercase
alphabetList = list(upperCaseString)
for i in range(0, 50):
    # this if statement will give us AA and AB etc up until AX since the grids will never be larger than this
    if i >= 26:
        numToLetter[i+1] = alphabetList[0] + alphabetList[i-26]
    else:
        numToLetter[i+1] = alphabetList[i]

# create border types for when we create new columns
border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
border_thick_bottom = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thick'))

# Function that will add the contents of each attendance file to the grid
def update_xlsx(src, dest):
    # open the xlxs file for reading (grid)
    wb = load_workbook(filename = dest)
    # get the current active worksheet (standard)
    ws = wb['Standard']
    # open the csv file
    with open(src) as fin:
        # We need to specify that these are global variables before we can edit them in this function
        global lastRow 
        global totalsRow 
        global unusedNames
        global synonym
        global notSynonym
        # read the csv
        reader = csv.reader(fin)
        # determine the day before finding the column using the csv file name
        dayNum = ''
        for i in src:
            if i.isdigit():
                dayNum = dayNum + i
            else:
                break
        # Find the current number of columns in the grid
        gridColumnCount = ws.max_column - 1
        
        # Determine which column the current csv file people need to go into in the grid based on the date.
        for i in range(3, gridColumnCount):
            # need to make sure that it is still recognized even if it has an r attached to it
            # so instead of comparing them directly, we can just check if the day is in the column title, since we go chronologically
            if str(dayNum) in str(ws.cell(row=2, column=i).value):
                #this means that the number at the top of the column is the same as the first letter of the csv file name, which is the date
                currentColumn = i
                # break so we stop at the first instance of that day
                break
        # check if it has a dash
        dash = 0
        # if it is a single digit day and it has a dash
        if str(src[1]) == '-':
            dash = 1
        # if it is a double digit day and it has a dash
        elif str(src[2]) == '-':
            dash = 2    
        # if the name has no dash
        else:
            dash = 0
        
        review = 0
        if str(src[1]).lower() == 'r': # if it is i single digit day ex: 1r
            review = 1
        elif str(src[2]).lower() == 'r': # if it is a double digit day ex: 11r
            review = 2
        elif str(src[3]).lower() == 'r': # if it is a single digit day with multiple sessions ex: 1-1r
            review = 3
        elif str(src[4]).lower() == 'r': # if it is a double digit day with multiple sessions ex: 11-1rn
            review = 4
        else:
            review = 0

        # if we are on a day with multiple sessions
        if dash != 0:
            # if it is the second session of the day
            if str(src[dash + 1]) == '2':
                # if it is the second session and there is already a second column, we just move into the second one.
                # this is to account for SI leaders just updating a session that already exists
                if str(ws.cell(row=2, column=currentColumn).value).replace('r', '') == str(ws.cell(row=2, column=currentColumn+1).value).replace('r', ''):
                    currentColumn += 1
                # if it is the second session of the day and there is only 1 column we go ahead and add the column after the current column
                elif str(ws.cell(row=2, column=currentColumn).value).replace('r', '') != str(ws.cell(row=2, column=currentColumn+1).value).replace('r', ''):     
                    # insert the new column
                    ws.insert_cols(currentColumn+1)
                    # move into that new column
                    currentColumn += 1
                    # give every cell in the column a border
                    for i in range(2, totalsRow):
                        # if it is the top row with the number, we want the border to have a thick bottom
                        if i == 2:
                            ws.cell(row=i, column=currentColumn).border = border_thick_bottom
                        # otherwise we give it a normal border
                        else:
                            ws.cell(row=i, column=currentColumn).border = border
                    # add the day at the top of the column, making sure to store it as an integer
                    ws.cell(row=2, column=currentColumn).value = int(str(ws.cell(row=2, column=currentColumn-1).value).replace('r', ''))
                    # make the top number bold and have the correct font size of 13
                    ws['{}{}'.format(numToLetter[currentColumn], 2)].font = Font(size = 13, bold=True)
                    # add the sum formulas to the 2 rows that store the totals on each column
                    # determine the letter of the current column
                    colLetter = numToLetter[currentColumn]
                    # sum from row 3 to right before 'names not listed'
                    ws['{}{}'.format(colLetter, namesNotListed)] = '=SUM({}{}:{}{})'.format(colLetter, 3, colLetter, namesNotListed-1)
                    # sum from 'names not listed' to the row with the last name on the list
                    ws['{}{}'.format(colLetter, totalsRow)] = '=SUM({}{}:{}{})'.format(colLetter, namesNotListed, colLetter, totalsRow -1)
                    # every time we add a new column, we make sure all columns have the correct sum formulas
                    updatedColCount = ws.max_column
                    for i in range(3, updatedColCount):
                        colLetter = numToLetter[i]
                        ws['{}{}'.format(colLetter, namesNotListed)] = '=SUM({}{}:{}{})'.format(colLetter, 3, colLetter, namesNotListed-1)
                        ws['{}{}'.format(colLetter, totalsRow)] = '=SUM({}{}:{}{})'.format(colLetter, namesNotListed, colLetter, totalsRow -1)
                    # update the last column that counts how many days had attendance
                    ws['{}{}'.format(numToLetter[updatedColCount], namesNotListed)] = '=COUNTIF({}{}:{}{},">0")'.format(numToLetter[3], totalsRow, numToLetter[updatedColCount-2], totalsRow)
            # if it is the third session 
            if str(src[dash + 1]) == '3':
                # if it is the third session and there is already a third column we just move into that third column
                if str(ws.cell(row=2, column=currentColumn).value).replace('r', '') == str(ws.cell(row=2, column=currentColumn+2).value).replace('r', ''):
                    currentColumn += 2
                # if it is not equal to the third column then check to see if it is equal to the second column
                elif str(ws.cell(row=2, column=currentColumn).value).replace('r', '') != str(ws.cell(row=2, column=currentColumn+2).value).replace('r', ''):
                    # check to see if it is equal to the second column
                    if str(ws.cell(row=2, column=currentColumn).value).replace('r', '') == str(ws.cell(row=2, column=currentColumn+1).value).replace('r', ''):
                        # if it is then we need to move into the second column and create a new column then move into it
                        currentColumn += 1
                        ws.insert_cols(currentColumn+1)
                        # move into the newly created column
                        currentColumn += 1
                        # give each cell in the column a border
                        for i in range(2, totalsRow):
                            # if it is the top row with the number, we want the border to have a thick bottom
                            if i == 2:
                                ws.cell(row=i, column=currentColumn).border = border_thick_bottom
                            else:
                                ws.cell(row=i, column=currentColumn).border = border
                        # add the day at the top of the column
                        ws.cell(row=2, column=currentColumn).value = int(str(ws.cell(row=2, column=currentColumn-1).value).replace('r', ''))
                        # make it bold
                        ws['{}{}'.format(numToLetter[currentColumn], 2)].font = Font(size = 13, bold=True)
                        # add the sum formula to the 2 totals rows = lastRow and totalsRow
                        # we first need to find which letter the current column corresponds to using our dictionary of letters 
                        colLetter = numToLetter[currentColumn]
                        # add up to names not listed
                        ws['{}{}'.format(colLetter, namesNotListed)] = '=SUM({}{}:{}{})'.format(colLetter, 3, colLetter, namesNotListed-1)
                        # add from names not listed to the last current name
                        ws['{}{}'.format(colLetter, totalsRow)] = '=SUM({}{}:{}{})'.format(colLetter, namesNotListed, colLetter, totalsRow - 1)
                        # janky but every time we add a column, we loop through and give all the columns the right formula
                        updatedColCount = ws.max_column
                        for i in range(3, updatedColCount):
                            # we first need to find which letter the current column corresponds to using our dictionary of letters 
                            colLetter = numToLetter[i]
                            # add up to names not listed
                            ws['{}{}'.format(colLetter, namesNotListed)] = '=SUM({}{}:{}{})'.format(colLetter, 3, colLetter, namesNotListed-1)
                            # add from names not listed to the last current name
                            ws['{}{}'.format(colLetter, totalsRow)] = '=SUM({}{}:{}{})'.format(colLetter, namesNotListed, colLetter, totalsRow -1)
                        # I also need to update that final column that adds up attendance from people in other sections
                        ws['{}{}'.format(numToLetter[updatedColCount], namesNotListed)] = '=COUNTIF({}{}:{}{},">0")'.format(numToLetter[3], totalsRow,numToLetter[updatedColCount-2], totalsRow)
                        
                    # if it not equal to the second column then we need to create 2 new columns and move into the third
                    else:
                        # create the two columns
                        for i in range(0, 2):
                            ws.insert_cols(currentColumn+1)
                            # move into the new column
                            currentColumn += 1
                            # give each cell in the column a border
                            for j in range(2, totalsRow):
                                # if it is the top row with the number, we want the border to have a thick bottom
                                if i == 2:
                                    ws.cell(row=i, column=currentColumn).border = border_thick_bottom
                                else:
                                    ws.cell(row=i, column=currentColumn).border = border
                            # add the day at the top of the column
                            ws.cell(row=2, column=currentColumn).value = int(str(ws.cell(row=2, column=currentColumn-1).value).replace('r', ''))
                            # make it bold
                            ws['{}{}'.format(numToLetter[currentColumn], 2)].font = Font(size = 13, bold=True)
                            # add the sum formula to the 2 totals rows = lastRow and totalsRow
                            # we first need to find which letter the current column corresponds to using our dictionary of letters 
                            colLetter = numToLetter[currentColumn]
                            # add up to names not listed
                            ws['{}{}'.format(colLetter, namesNotListed)] = '=SUM({}{}:{}{})'.format(colLetter, 3, colLetter, namesNotListed-1)
                            # add from names not listed to the last current name
                            ws['{}{}'.format(colLetter, totalsRow)] = '=SUM({}{}:{}{})'.format(colLetter, namesNotListed, colLetter, totalsRow -1)
                            # janky but every time we add a column, we loop through and give all the columns the right formula
                            updatedColCount = ws.max_column
                            for i in range(3, updatedColCount):
                                # we first need to find which letter the current column corresponds to using our dictionary of letters 
                                colLetter = numToLetter[i]
                                # add up to names not listed
                                ws['{}{}'.format(colLetter, namesNotListed)] = '=SUM({}{}:{}{})'.format(colLetter, 3, colLetter, namesNotListed-1)
                                # add from names not listed to the last current name
                                ws['{}{}'.format(colLetter, totalsRow)] = '=SUM({}{}:{}{})'.format(colLetter, namesNotListed, colLetter, totalsRow -1)
                            # I also need to update that final column that adds up attendance from people in other sections
                            ws['{}{}'.format(numToLetter[updatedColCount], namesNotListed)] = '=COUNTIF({}{}:{}{},">0")'.format(numToLetter[3], totalsRow, numToLetter[updatedColCount-2], totalsRow)
                        # We should now be in the third column so we are good to go
        
        # if the current day is a review day we add an 'r' to the top of the column
        if review != 0:
            # add the r to the column
            ws.cell(row=2, column=currentColumn).value = 'r' + dayNum

        # make all of the columns the same width
        for i in range(3, ws.max_column):
            ws.column_dimensions['{}'.format(numToLetter[i])].width = 5

        # enumerate the rows so that we can access individual names and last names
        for index,row in enumerate(reader):
            #skip the first 2 rows since they just have the column header and my name
            if index == 0 or index == 1:
                continue
            # Since the first and last names are separated by space, we can split it into 2 different strings. row[0] = first name & row[1] = last name
            row = row[0].split()
            # if the cell is blank we don't want to do anything with it
            if len(row) == 0:
                continue
            # if they just put a first name check to see if maybe they put their last name initial in caps at the end of their name
            if len(row) == 1:
                nameLen = len(row[0])
                initialCheck = 0
                if nameLen > 1:
                    for j in range(0, nameLen):
                        # if the last letter is uppercase
                        if row[0][j].isupper() and j == nameLen - 1:
                            # we can go ahead and store that letter and make it the last name initial and remove it from first name
                            initial = str(row[0][j])
                            row[0] = row[0][:-1]
                            initialCheck = 1
                            row.append(initial)
                            break
                if initialCheck == 0:
                    if src[1] == '.' or src[1] == '-' or src[1] == 'r':
                        unusedNames.append("Day of month: " + src[0])
                    else:
                        unusedNames.append("Day of month: " + src[0] + src[1])
                    unusedNames.append("Name: " + str(row))
                    continue
            # if more than 2 names total we will only use the first one as the first name and the last one as the last name
            if len(row) > 2:
                lenNames = len(row)
                # store the first and the last name only
                row[1] = str(row[lenNames - 1])
            # if the last name is only 2 letters
            if len(row) > 1 and len(row[1]) == 2:
                # check to see if the last name is on the grid
                isInGrid = 0
                for i in range(3, lastRow):
                    # if the last name is in the grid
                    if str(row[1]).lower() == str(ws.cell(row=i, column=1).value).lower():
                        # if the first name matches then we let it through to get added properly
                        if str(row[0]).lower == str(ws.cell(row=i, column=2).value).lower(): 
                            isInGrid = 1
                # if the last name is not in the grid we can skip it and just add it to the list of names that didn't make it
                if not isInGrid:
                    if src[1] == '.' or src[1] == '-' or src[1] == 'r':
                        unusedNames.append("Day of month: " + src[0])
                    else:
                        unusedNames.append("Day of month: " + src[0] + src[1])
                    unusedNames.append("Name: " + str(row))
                    continue;
            # if they were observed and the first name is in the mentors dictionary and the value is the same as the last name
            if numObservations != 0 and str(row[0]).lower() in mentors and mentors[str(row[0]).lower()] == str(row[1]).lower():
                # do not put a mentor name in the grid
                continue
            # check will change to 1 if the name is found so that we don't add any duplicate names
            check = 0
            reportLast = str(row[1]).lower()
            reportFirst = str(row[0]).lower()
            # remove any symbols from the names of the attendance reports
            for char in string.punctuation:
                reportFirst = reportFirst.replace(char, '')
                reportLast = reportLast.replace(char, '')
            # Compare each name and last name from the csv file to the grid to add the '1' in the appropriate cell
            for i in range(3, lastRow):
                # store the first and last name on each row of the grid into a temp vari
                gridLast = str(ws.cell(row=i, column=1).value).lower()
                gridFirst = str(ws.cell(row=i, column=2).value).lower()
                # remove any symbols from both last names
                for char in string.punctuation:
                    gridFirst = gridFirst.replace(char, '')
                    gridLast = gridLast.replace(char, '')
                # loop for however many last names there are for that one person on the grid
                tempArr = gridLast.split()
                for j in range(0, len(tempArr)):
                    # check if the last name matches, even if the person just put in an initial 
                    if reportLast == tempArr[j] or (len(reportLast) == 1 and reportLast == tempArr[j][0] and reportFirst[0] == gridFirst[0]) or (len(reportLast) > 1 and reportLast in tempArr[j]) or gridLast.replace(' ', '') == reportLast:
                        # check for however many first names there are if they match
                        tempFArr = gridFirst.split()
                        for k in range(0, len(tempFArr)):
                            if reportFirst == tempFArr[k] or (len(reportFirst) == 1 and reportFirst == tempFArr[k][0]):
                                # add the 1 to the right cell
                                ws.cell(row=i, column=currentColumn).value = 1
                                # change check to 1 to show that we found the name
                                check = 1
                                # break so we don't keep trying to find a name we already found
                                break
                        # if the first name does not match we want to make sure it isn't just a nickname
                        if check == 0 and (gridFirst[0] == reportFirst[0] or reportFirst in gridFirst): 
                            # check if the key is in the dictionary and then check if the value of the key is the same as the last name on the grid
                            if str(row[0]).lower() in synonym and synonym[str(row[0]).lower()] == str(ws.cell(row=i, column=2).value).lower():
                                # if the value of the key is the same as the last name on the grid then we don't have to ask and we can just add the 1 to the right place
                                # add the 1 to the right cell
                                ws.cell(row=i, column=currentColumn).value = 1
                                # change check to 1 to show that we found the name
                                check = 1
                                # break so we don't keep trying to find a name we already found
                                break
                            # check if the key is in the notSynonym dictionary and check that the value stored for that key is the same the last name on the grid. if so continue.
                            elif str(row[0]).lower() in notSynonym and notSynonym[str(row[0]).lower()] == str(ws.cell(row=i, column=2).value).lower():
                                continue
                            else:
                                # ask the user to see if the person used a nickname, perhaps. Ex: andy instead of andrew
                                response = input('Is {} {} the same as {} {}? Y or N: '.format(str(row[0]).lower(), reportLast, str(ws.cell(row=i, column=2).value).lower(), gridLast))
                                #print('\nIs', str(row[0]).lower(), ' the same as ', str(ws.cell(row=i, column=2).value).lower(), '? Y or N: ')
                                #response = input()
                                if str(response).lower() == 'y':
                                    # add the 1 to the right cell
                                    ws.cell(row=i, column=currentColumn).value = 1
                                    # change check to 1 to show that we found the name
                                    check = 1
                                    # add the new value to the dictionary of synonyms
                                    synonym[str(row[0]).lower()] = str(ws.cell(row=i, column=2).value).lower()
                                    # break so we don't keep trying to find a name we already found
                                    break
                                # if the user says that these are not the same people then we add the names to notSynonym
                                else:
                                    # add the 2 names so that we don't ask again
                                    notSynonym[str(row[0]).lower()] = str(ws.cell(row=i, column=2).value).lower()
                                    continue
            # if we did not find the name (check = 0) then we add it to the next open row, which was given as a global variable lastRow
            if check == 0 and lastRow != totalsRow:
                ws.cell(row=lastRow, column=1).value = reportLast # add the last name to the first column
                ws.cell(row=lastRow, column=2).value = reportFirst # add the first name to the second column
                ws.cell(row=lastRow, column=currentColumn).value = 1  # add a '1' to the appropriate cell to show attendance from this person
                lastRow += 1 # increase the lastRow to show that the next open row is now 1 below this new name
            # if we did not find the name but we are on the row with the totals, we don't want to overwrite any of those, so we add a row before lastRow before adding the name
            elif check == 0 and lastRow == totalsRow:
                ws.insert_rows(lastRow) # add a new row right before the limit
                ws.cell(row=lastRow, column=1).value = reportLast # add the last name 
                ws.cell(row=lastRow, column=2).value = reportFirst # add the first name
                ws.cell(row=lastRow, column=currentColumn).value = 1 # add a '1' to the appropriate cell to show attendance from this person
                lastRow += 1 # increment the last row value by one
                totalsRow += 1 # increment the totals row value by 1
                
    # save changes made to the grid
    wb.save(dest)
            
            
# Enter the name of the grid along with its file type. ex: grids_etc.xlsx
gridName = input("\nName of grid file: (don't forget .xlsx)\nEx: COP_3502C_SG_February.xlsx\n\n")
# How many days do you want to input
total = int(input('\nNumber of attendance reports: '))

fileNameArr = [] # this makes fileNameArr a list
print('\nEnter attendance report names: (Press enter between each one)')
# Scan each day one at a time and append the .csv file to each one
for i in range(0, total):
    temp = str(input())
    # if the file is a review file we will add it to the front to make sure that it gets added first
    fileNameArr.append(temp + '.csv')
print('')

# So we can call our function to run for each file and update the grid
for i in range(0, total):
    update_xlsx(fileNameArr[i], gridName)

# function to find the right columns to fill and fills them with the color passed in
def fillColumn(file_name, daysArr, numDays, color):
    wb = load_workbook(filename=file_name)
    ws = wb['Standard']
    colorFill = PatternFill(start_color=color, fill_type='solid')
    numColumns = ws.max_column - 1

    for i in range(0, numDays):
        for j in range(3, numColumns):
            # if the column number matches with the day we want to fill
            if str(daysArr[i]) == str(ws.cell(row=2, column=j).value):
                # fill the entire column with the color
                for k in range(2, totalsRow):
                    ws.cell(row=k, column=j).fill = colorFill
                wb.save(file_name)
                # don't break because we want all of them to be red

# this function will find a letter in the 
def fillCol(file_name, letter, color):
    wb = load_workbook(filename=file_name)
    ws = wb['Standard']
    colorFill = PatternFill(start_color=color, fill_type='solid')
    numColumns = ws.max_column - 1    

    for i in range(3, numColumns):
        # check if the column has an r on it 
        val = str(ws.cell(row=2, column=i).value)
        if letter in val:
            # if it has an r we remove the r and fill it with green
            val = val.replace(letter, '')
            # put the right number at the top of the column
            ws.cell(row=2, column=i).value = int(val)
            # bold that number
            ws['{}{}'.format(numToLetter[i], 2)].font = Font(size = 13, bold=True)
            for j in range(2, totalsRow):
                ws.cell(row=j, column=i).fill = colorFill
            wb.save(file_name)


# fill in red on test days
numExams = int(input("\nNumber of exams this month: "))
exams = []

# get all of the exam dates
for i in range(0, numExams):
    examTemp = int(input("Exam day #{}: ".format(i + 1)))
    exams.append(examTemp)

# fill in the entire column with red
fillColumn(gridName, exams, numExams, 'FFFF0000')

# fill the exam days green
fillCol(gridName, 'r', '00FF00')

# the commented code below was the old way of coloring reviews green
'''
# fill in green on review days
numReviews = int(input("\nHow many exam reviews did you hold this month? "))
reviews = []

# get all of the review dates

for i in range(0, numReviews):
    revTemp = int(input("Review day #{}: ".format(i + 1)))
    reviews.append(revTemp)

# fill in the entire column with green
fillColumn(gridName, reviews, numReviews, '00FF00')
'''

# Print out the names that were not used so the SI Leader can do what they want with the information
if len(unusedNames) != 0:
    print("\n================================================================================")
    print('\nThe names below were not added to the grid due to formatting issue, such as: ')
    print('\t1) Being just a first name.\n\t2) Only having 2 letters in their last name.')
    print('\tFeel free to go back and add their attendance using your own discretion.\n')
    # for loop to print more neatly
    # determine how many lines to print in case of only 2 names
    for i in range(0, int(len(unusedNames) / 2) + 2):
        # since we print in pairs we skip any time i is odd 
        if i % 2 == 0:
            print('\t- ', unusedNames[i], ' ', unusedNames[i + 1], '\n')
    print("================================================================================")
# Ask the user if they are done with the program
finish = input('\nPRESS ANY LETTER + ENTER TO EXIT THE PROGRAM.\n')