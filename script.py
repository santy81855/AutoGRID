from openpyxl import load_workbook # for opening xlsx files like the grids
from openpyxl.styles import PatternFill # to fill columns
import csv # for opening csv  files like the attendance reports

# create global variable to store the first open line in case we have to add any names to the grid
lastRow = int(input('\nWhat is the number of the first blank row on your grid? This refers to the first row we could add a name to: '))
# create global variable to store the line which has the totals at the bottom so we don't overwrite it
totalsRow = int(input('What is the number of the row that has all of the totals on it? This row number should be larger than the row number you put above: '))
# create global variable to store the names of any mentors that may have observed the SI leader so we don't put their names in the grid
mentors = {}
# ask how many times they were observed
numObservations = int(input("How many times were you observed this month? "))
if numObservations != 0:
    print('\n')
for i in range(0, numObservations):
    temp1 = str(input("First name of mentor #{}: ".format(i+1)))
    temp2 = str(input("Last name of mentor #{}: ".format(i+1)))
    # add the first name of the mentor as a key in the dictionary and the last name as the value
    mentors[temp1] = temp2
# create a global variable that will be a list that will store all of the names that were not used along with the days they attended so the user can do what they want with those
unusedNames = []
# create global dictionary to store synonyms
synonym = {}
# create global dictionary to store non synonyms
notSynonym = {}

# Function that will add the contents of each attendance file to the grid
def update_xlsx(src, dest):
    # open the xlxs file for reading (grid)
    wb = load_workbook(filename = dest)
    # get the current active worksheet (standard)
    ws = wb['Standard']
    # open the csv file
    with open(src) as fin:
        # read the csv
        reader = csv.reader(fin)
        # Find the number of columns in the grid
        gridColumnCount = ws.max_column - 1
        # Determine which column the current csv file people need to go into in the grid based on the date.
        # This is done by comparing the first character of each file name ex: 1.csv to the numbers at the top of each column in the grid.
        # We also compare the first two characters of the csv file in case it is a double digit date.
        for i in range(3, gridColumnCount):
            if str(src[0]) == str(ws.cell(row=2, column=i).value) or str(src[0]+src[1]) == str(ws.cell(row=2, column=i).value):
                #this means that the number at the top of the column is the same as the first letter of the csv file name, which is the date
                currentColumn = i;
        global lastRow # need to specify it is a global variable if I want to modify it 
        global totalsRow # same thing
        global unusedNames # same thing
        global synonym
        global notSynonym
        # enumerate the rows so that we can access individual names and last names
        for index,row in enumerate(reader):
            #skip the first 2 rows since they just have the column header and my name
            if index == 0 or index == 1:
                continue
            # Since the first and last names are separated by space, we can split it into 2 different strings. row[0] = first name & row[1] = last name
            row = row[0].split()
            
            # if the cell is blank we don't want to do anything with it
            if len(row) == 0:
                continue
            # ignore people who only put a first name
            if len(row) == 1:
                if src[1] == '.':
                    unusedNames.append("Day of month: " + src[0])
                else:
                    unusedNames.append("Day of month: " + src[0] + src[1])
                unusedNames.append("Name: " + str(row))
                continue
            # ignore people who have more than 2 names
            if len(row) > 2:
                if src[1] == '.':
                    unusedNames.append("Day of month: " + src[0])
                else:
                    unusedNames.append("Day of month: " + src[0] + src[1])
                unusedNames.append("Name: " + str(row))
                continue
            # if they put a single letter as their last name
            if len(row[1]) == 1:
                if src[1] == '.':
                    unusedNames.append("Day of month: " + src[0])
                else:
                    unusedNames.append("Day of month: " + src[0] + src[1])
                unusedNames.append("Name: " + str(row))
                continue
            # if they put a single letter as they first name
            if len(row[0]) == 1:
                if src[1] == '.':
                    unusedNames.append("Day of month: " + src[0])
                else:
                    unusedNames.append("Day of month: " + src[0] + src[1])
                unusedNames.append("Name: " + str(row))
                continue;
            # if they put 2 letters as their last name
            if len(row[1]) == 2:
                if src[1] == '.':
                    unusedNames.append("Day of month: " + src[0])
                else:
                    unusedNames.append("Day of month: " + src[0] + src[1])
                unusedNames.append("Name: " + str(row))
                continue;
            # if they were observed and the first name is in the mentors dictionary and the value is the same as the last name
            if numObservations != 0 and str(row[0]).lower() in mentors and mentors[str(row[0]).lower()] == str(row[1]).lower():
                # do not put a mentor name in the grid
                continue
            # check will change to 1 if the name is found so that we don't add any duplicate names
            check = 0
            # Compare each name and last name from the csv file to the grid to add the '1' in the appropriate cell
            for i in range(3, lastRow):
                # if the last name matches we can then check if the first name matches as well
                if str(row[1]).lower() == str(ws.cell(row=i, column=1).value).lower():
                    # check if the first name matches and add the 1 to the right cell
                    if str(row[0]).lower() == str(ws.cell(row=i, column=2).value).lower():
                        # add the 1 to the right cell
                        ws.cell(row=i, column=currentColumn).value = 1
                        # change check to 1 to show that we found the name
                        check = 1
                        # break so we don't keep trying to find a name we already found
                        break
                    # if the first name does not match
                    else:
                        # check if the key is in the dictionary and then check if the value of the key is the same as the last name on the grid
                        if str(row[0]).lower() in synonym and synonym[str(row[0]).lower()] == str(ws.cell(row=i, column=2).value).lower():
                            # if the value of the key is the same as the last name on the grid then we don't have to ask and we can just add the 1 to the right place
                            # add the 1 to the right cell
                            ws.cell(row=i, column=currentColumn).value = 1
                            # change check to 1 to show that we found the name
                            check = 1
                            # break so we don't keep trying to find a name we already found
                            break
                        # check if the key is in the notSynonym dictionary and check that the value stored for that key is the same the last name on the grid. if so continue.
                        elif str(row[0]).lower() in notSynonym and notSynonym[str(row[0]).lower()] == str(ws.cell(row=i, column=2).value).lower():
                            continue
                        else:
                            # ask the user to see if the person used a nickname, perhaps. Ex: andy instead of andrew
                            response = input('Is {} the same as {}? Y or N: '.format(str(row[0]).lower(), str(ws.cell(row=i, column=2).value).lower()))
                            #print('\nIs', str(row[0]).lower(), ' the same as ', str(ws.cell(row=i, column=2).value).lower(), '? Y or N: ')
                            #response = input()
                            if str(response).lower() == 'y':
                                # add the 1 to the right cell
                                ws.cell(row=i, column=currentColumn).value = 1
                                # change check to 1 to show that we found the name
                                check = 1
                                # add the new value to the dictionary of synonyms
                                synonym[str(row[0]).lower()] = str(ws.cell(row=i, column=2).value).lower()
                                # break so we don't keep trying to find a name we already found
                                break
                            # if the user says that these are not the same people then we add the names to notSynonym
                            else:
                                # add the 2 names so that we don't ask again
                                notSynonym[str(row[0]).lower()] = str(ws.cell(row=i, column=2).value).lower()
                                continue
                   
            # if we did not find the name (check = 0) then we add it to the next open row, which was given as a global variable lastRow
            if check == 0 and lastRow != totalsRow:
                # we will only add a last name if both a first and last name are given. We do this to avoid crashing if they just input a first name
                if len(row) == 2:
                    ws.cell(row=lastRow, column=1).value = str(row[1]) # add the last name to the first column
                ws.cell(row=lastRow, column=2).value = str(row[0]) # add the first name to the second column
                ws.cell(row=lastRow, column=currentColumn).value = 1  # add a '1' to the appropriate cell to show attendance from this person
                lastRow += 1 # increase the lastRow to show that the next open row is now 1 below this new name
            # if we did not find the name but we are on the row with the totals, we don't want to overwrite any of those, so we add a row before lastRow before adding the name
            elif check == 0 and lastRow == totalsRow:
                ws.insert_rows(lastRow) # add a new row right before the limit
                # only add the last name if botha first and last name are given.
                if len(row) == 2:
                    ws.cell(row=lastRow, column=1).value = str(row[1]) # add the last name 
                ws.cell(row=lastRow, column=2).value = str(row[0]) # add the first name
                ws.cell(row=lastRow, column=currentColumn).value = 1 # add a '1' to the appropriate cell to show attendance from this person
                lastRow += 1 # increment the last row value by one
                totalsRow += 1 # increment the totals row value by 1
                
    # save changes made to the grid
    wb.save(dest)
            
            
# Enter the name of the grid along with its file type. ex: grids_etc.xlsx
gridName = input('\nPlease enter the name of the grid file making sure to include the .xlsx file type.\nEx: COP_3502C_SG_February.xlsx\n')
# How many days do you want to input
total = int(input('\nHow many days of attendance do you want to input? '))

fileNameArr = [] # this makes fileNameArr a list
print('\nEnter only the days of the month. Enter one at a time, pressing enter after each day. Ex: 8 9 15 16')
# Scan each day one at a time and append the .csv file to each one
for i in range(0, total):
    temp = input()
    fileNameArr.append(temp + '.csv')

# So we can call our function to run for each file and update the grid
for i in range(0, total):
    update_xlsx(fileNameArr[i], gridName)

# function to find the right columns to fill and fills them with the color passed in
def fillColumn(file_name, daysArr, numDays, color):
    wb = load_workbook(filename=file_name)
    ws = wb['Standard']
    colorFill = PatternFill(start_color=color, fill_type='solid')
    numColumns = ws.max_column - 1

    for i in range(0, numDays):
        for j in range(3, numColumns):
            # if the column number matches with the day we want to fill
            if int(daysArr[i]) == int(ws.cell(row=2, column=j).value):
                # fill the entire column with the color
                for k in range(2, totalsRow):
                    ws.cell(row=k, column=j).fill = colorFill
                wb.save(file_name)
                break

# fill in red on test days
numExams = int(input("\nHow many exams did the class have this month? "))
exams = []

# get all of the exam dates
for i in range(0, numExams):
    examTemp = int(input("Exam day #{}: ".format(i + 1)))
    exams.append(examTemp)

# fill in the entire column with red
fillColumn(gridName, exams, numExams, 'FFFF0000')

# fill in green on review days
numReviews = int(input("\nHow many reviews did you hold this month? "))
reviews = []

# get all of the review dates
for i in range(0, numReviews):
    revTemp = int(input("Review day #{}: ".format(i + 1)))
    reviews.append(revTemp)

# fill in the entire column with green
fillColumn(gridName, reviews, numReviews, '00FF00')

# Print out the names that were not used so the SI Leader can do what they want with the information
if len(unusedNames) != 0
    print('\n\nThe names below were not accounted for due to one of the following reasons:')
    print('\t1) Only having a first name\n\t2) Having more than 2 names total\n\t3) Using initials for either first or last name\n')
    print('Feel free to go back and add their attendance using your own discretion\n')
    # for loop to print more neatly
    for i in range(0, int(len(unusedNames) / 2)):
        # since we print in pairs we skip any time i is odd 
        if i % 2 == 0:
            print(unusedNames[i], ' ', unusedNames[i + 1], '\n')