import sys
import os
from PyQt5.uic import loadUi
from PyQt5 import QtWidgets, QtCore, QtGui
from PyQt5.QtWidgets import QDialog, QApplication, QWidget, QFileDialog, QInputDialog, QLineEdit, QPushButton
from openpyxl import load_workbook # for opening xlsx files like the grids
from openpyxl.styles import PatternFill # to fill columns
from openpyxl.styles.borders import Border, Side # to create borders an new columns
from openpyxl.styles import Font # to write the bold number at the top of columns
import csv # for opening csv  files like the attendance reports
import string # for removing symbols from names
import ctypes
from datetime import datetime
import xlrd
from openpyxl.styles.alignment import Alignment
import sys
# to get the working monitor size
from win32api import GetMonitorInfo, MonitorFromPoint
from PyQt5 import QtCore
from PyQt5 import QtGui
from PyQt5.QtGui import QCursor, QMouseEvent, QFont, QKeySequence, QSyntaxHighlighter, QTextCharFormat, QBrush, QTextCursor
from PyQt5.QtCore import QPoint, pyqtSignal, QRegExp
from PyQt5.QtCore import Qt, QPropertyAnimation, QRect, QEasingCurve
from PyQt5.QtCore import QObject, QMimeData
from PyQt5.QtWidgets import QApplication, QMainWindow, QLineEdit, QCompleter, QFileDialog, QGraphicsDropShadowEffect
from PyQt5.QtWidgets import QHBoxLayout, QTextEdit, QPlainTextEdit, QShortcut, QScrollArea
from PyQt5.QtWidgets import QLabel, QStackedWidget, QMessageBox
from PyQt5.QtWidgets import QPushButton, QDesktopWidget
from PyQt5.QtWidgets import QVBoxLayout, QScrollBar
from PyQt5.QtWidgets import QWidget, QFrame
from PyQt5.QtCore import Qt, QRect, QSize, QRectF
from PyQt5.QtWidgets import QWidget, QPlainTextEdit, QTextEdit
from PyQt5.QtGui import QColor, QPainter, QTextFormat, QLinearGradient
import os
import ctypes

import TitleBar, FirstWindow, config, AutoGrid

# globals
# this sets the icon as your taskbar icon
myappid = 'AutoGrid'
ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID(myappid)

# The code below makes it so that the widget doesn't get messed up if you scale windows text by 125% or more
# Query DPI Awareness (Windows 10 and 8)
awareness = ctypes.c_int()
errorCode = ctypes.windll.shcore.GetProcessDpiAwareness(0, ctypes.byref(awareness))
# Set DPI Awareness  (Windows 10 and 8)
errorCode = ctypes.windll.shcore.SetProcessDpiAwareness(0)
# the argument is the awareness level, which can be 0, 1 or 2:
# for 1-to-1 pixel control I seem to need it to be non-zero (I'm using level 2)
# Set DPI Awareness  (Windows 7 and Vista)
success = ctypes.windll.user32.SetProcessDPIAware()
# behaviour on later OSes is undefined, although when I run it on my Windows 10 machine, it seems to work with effects identical to SetProcessDpiAwareness(1)

# create the global variables
current_month = 0
grid_name = ''
attendance_sheet_name = ''
first_blank_row = 0
names_not_listed_row = 0
totals_row = 0
num_observations = 0
num_exams = 0
exam_days = []
mentors = {}
synonym = {}
notSynonym = {}
done_observations = 0
isSame = '' # used to determine nicknames
zoom_attendance_reports = []
num_zoom_attendance_reports = 0
zoom_days = []
inPerson = False # used to determine if they have any inPerson attendance to report
isUsage = True
# create a dictionary that stores weekdays
weekdays = {}
weekdays[0] = "monday"
weekdays[1] = "tuesday"
weekdays[2] = "wednesday"
weekdays[3] = "thurssday"
weekdays[4] = "friday"
weekdays[5] = "saturday"
weekdays[6] = "sunday"

# create a modality dictionary that just assigns each modality a number
modality = {}
modality[0] = 'P'
modality[1] = 'R'
modality[2] = 'R'
modality[3] = 'HP'
modality[4] = 'HR'
modality[5] = 'HP'
modality[6] = 'HR'
modality[7] = 'RP'
modality[8] = 'RR'

# create a dictionary to take you from the first column of a day to the correct one based on what
# modality you need
correct_column = {}
correct_column['P1'] = 0
correct_column['P2'] = 1
correct_column['R1'] = 2
correct_column['R2'] = 3
correct_column['HP1'] = 4
correct_column['HR1'] = 5
correct_column['HP2'] = 6
correct_column['HR2'] = 7
correct_column['RP'] = 8
correct_column['RR'] = 9

# create a dictionary for each day of the month that tracks what type of sessions are happening on
# each day
class Session:
    # two in person sessions in one day (P) 
    # this implies it is not a hybrid session
    person1 = False 
    person2 = False
    # two possible remote sessions in one day (R)
    # this implies it is not a hybrid session
    remote1 = False
    remote2 = False
    # two possible hybrid sessions in one day (HR & HP)
    # this would take up to four columns
    hybrid1 = False
    hybrid2 = False
    # review which could be P, R, or HP & HR
    # if both are true for the same day then it was a hybrid review
    review_person = False
    review_remote = False
    # variable to track if this day has already been created    
    created = False

# create dictionary storing the 31 days of the month
session_types = {}
for i in range(1, 32):
    session_types[i] = Session()

# use a dictionary to keep track of what days were used
used_days = {}
for i in range(1, 32):
    used_days[i] = False

# create dictionary of letters and numbers so we can find the proper letter for a certain column number
numToLetter = {} # numbers are keys and letters are values
# create string of uppercase letters
upperCaseString = string.ascii_uppercase
alphabetList = list(upperCaseString)
'''
for i in range(0, 50):
    # this if statement will give us AA and AB etc up until AX since the grids will never be larger than this
    if i >= 26:
        numToLetter[i+1] = alphabetList[0] + alphabetList[i-26]
    else:
        numToLetter[i+1] = alphabetList[i]
'''
for i in range(0, 200):
    numToLetter[i+1] = xlrd.formula.colname(i)

# create border types for when we create new columns
border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
border_thick_bottom = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thick'))

def runAutoGridZoom():
    # open the xlsx grid file
    wb = load_workbook(grid_name)
    # get the current active worksheet (The only one)
    ws = wb.worksheets[0]

    # we need to add each name from each zoom file to the grid
    for i in range(0, num_zoom_attendance_reports):
        runAutoGridZoomHelper(wb, ws, zoom_attendance_reports[i], zoom_days[i])
    
    # go through all the columns and remove the ones that we don't need
    updatedColCount = ws.max_column
    temp_day = 0
    i = 4
    
    while str(ws.cell(row=2,column=i).value) != 'None':
        # store the day at the top of the column in temp_day
        temp_day = int(ws.cell(row=2,column=i).value)
        # check if the day was even used or if we can skip it
        if used_days[temp_day] == False:
            i += 1
            continue
        # if we did use it then we need to go col by col and delete any unused ones
        # if the first in-person session was not used then delete it
        if session_types[temp_day].person1 == False and str(ws.cell(row=1,column=i).value) == 'P':
            ws.delete_cols(i, 1)
        # if it was used then we need to move into the next column manually since we didn't delete
        else:
            i += 1
        # check if the second in-person wasn't used
        if session_types[temp_day].person2 == False and str(ws.cell(row=1,column=i).value) == 'P':
            ws.delete_cols(i, 1)
        else:
            i += 1
        
        # check if first remote session was used
        if session_types[temp_day].remote1 == False and str(ws.cell(row=1,column=i).value) == 'R':
            ws.delete_cols(i, 1)
        else:
            i += 1        
        # check if second remote was used
        if session_types[temp_day].remote2 == False and str(ws.cell(row=1,column=i).value) == 'R':
            ws.delete_cols(i, 1)
        else:
            i += 1    
        # check if first hybrid was used
        if session_types[temp_day].hybrid1 == False and str(ws.cell(row=1,column=i).value) == 'HP':
            ws.delete_cols(i, 2)
        else:
            i += 2   
        # check if second hybrid was used
        if session_types[temp_day].hybrid2 == False and str(ws.cell(row=1,column=i).value) == 'HP':
            ws.delete_cols(i, 2)
        else:
            i += 2    
        # check if person_review was used
        if session_types[temp_day].review_person == False and str(ws.cell(row=1,column=i).value) == 'RP':
            ws.delete_cols(i, 1)
        else:
            i += 1    
        # check if remote _review was used
        if session_types[temp_day].review_remote == False and str(ws.cell(row=1,column=i).value) == 'RR':
            ws.delete_cols(i, 1)
        else:
            i += 1    
    
    # add all the exam colorings necessary
    addExamColor(ws, wb, 'FFD92906', len(exam_days), exam_days)
    
    
    # every time we add a new column, we make sure all columns have the correct sum formulas
    updatedColCount = ws.max_column
    for i in range(4, updatedColCount):
        colLetter = numToLetter[i]
        ws['{}{}'.format(colLetter, names_not_listed_row)] = '=SUM({}{}:{}{})'.format(colLetter, 3, colLetter, names_not_listed_row-1)
        ws['{}{}'.format(colLetter, totals_row)] = '=SUM({}{}:{}{})'.format(colLetter, names_not_listed_row, colLetter, totals_row -1)
        # make sure all the columns have borders too
        for j in range(3, totals_row+1):
            ws.cell(row=j, column=i).border = border
    
    # make sure every name is centered and all names have borders
    for i in range(1, 4):
        for j in range(names_not_listed_row + 1, totals_row):
            # if the section number doesn't say N/A then add it
            if str(ws.cell(row=j, column=i).value) == "None":
                ws.cell(row=j, column=i).value = 'N/A'
            ws.cell(row=j, column=i).border = border
            ws.cell(row=j, column=i).alignment = Alignment(horizontal="center")
    
    #######################################################################################################################################
    # make sure all the columns have the correct sum formulas
    updatedColCount = ws.max_column
    for i in range(3, totals_row):
        if i == names_not_listed_row:
            continue
        colLetter = numToLetter[updatedColCount - 1]
        ws['{}{}'.format(colLetter, i)] = '=SUM({}{}:{}{})'.format(numToLetter[4], i, numToLetter[updatedColCount - 2], i)
    # update the last column that counts how many days had attendance
    updatedColCount = ws.max_column
    ws['{}{}'.format(numToLetter[updatedColCount], names_not_listed_row)] = '=COUNTIF({}{}:{}{},">0")'.format(numToLetter[4], totals_row, numToLetter[updatedColCount-2], totals_row)  
    #######################################################################################################################################
    
    # save because we are done entering names
    wb.save(grid_name)

def runAutoGridZoomHelper(wb, ws, attendance_report, file_day):
    global first_blank_row
    global names_not_listed_row
    global totals_row
    global isUsage
    global session_types

    # find the names_not_listed_row and first_blank_row and totals_row each time we want to add someone
    for i in range(2, ws.max_row + 1):
        temp = str(ws.cell(row=i, column=1).value)

        if 'not listed' in temp:
            names_not_listed_row = i

        if temp == 'None':
            first_blank_row = i
            break
    totals_row = ws.max_row

    with open(attendance_report) as fin:
        # read the csv
        reader = csv.reader(fin)
        # get the day out of the name and determine if it is double or not
        isDouble = 0
        # determine if it is a review or not
        isReview = 0
        # determine if it is a hybrid session or not
        isHybrid = 0
        length = len(file_day)
        if '-1' in file_day:
            isDouble = 1
        if '-2' in file_day:
            isDouble = 2
        if 'r' in file_day:
            isReview = True
        if 'h' in file_day:
            isHybrid = 1
        # day that is not a review or hybrid
        if 'r' not in file_day and 'h' not in file_day and (length == 4 or length == 2):
            cur_day = file_day[0] + file_day[1]
        elif 'r' not in file_day and 'h' not in file_day and (length == 3 or length == 1):
            cur_day = file_day[0]
        elif 'r' in file_day and length == 2:
            cur_day = file_day[0]
        elif 'r' in file_day and length == 3:
            cur_day = file_day[0] + file_day[1]
        elif 'h' in file_day and length == 2:
            cur_day = file_day[0]
        elif 'h' in file_day and length == 3:
            cur_day = file_day[0] + file_day[1]
        elif 'h' in file_day and length == 4:
            cur_day = file_day[0]
        elif 'h' in file_day and length == 5:
            cur_day = file_day[0] + file_day[1]
        # mark this day's object with whatever applies
        # it is a zoom review
        if isReview == True:
            session_types[int(cur_day)].review_remote = True
        else:
            # it is a regular zoom session
            if isDouble == 0 and isHybrid == False:
                session_types[int(cur_day)].remote1 = True
            # it is the first of 2 regular zoom sessions
            if isDouble == 1 and isHybrid == False:
                session_types[int(cur_day)].remote1 = True
                # make the second one true as well even if it doesn't have attendance
                session_types[int(cur_day)].remote2 = True
            # it is the second of 2 regular zoom sessions
            if isDouble == 2 and isHybrid == False:
                session_types[int(cur_day)].remote2 = True
                session_types[int(cur_day)].remote1 = True
            # it is the only hybrid session of the day
            if isHybrid == True and isDouble == 0:
                session_types[int(cur_day)].hybrid1 = True
            # it is the first of 2 hybrid sessions
            if isHybrid == True and isDouble == 1:
                session_types[int(cur_day)].hybrid1 = True
                session_types[int(cur_day)].hybrid2 = True
            # it is the second of 2 hybrid sessions
            if isHybrid == True and isDouble == 2:
                session_types[int(cur_day)].hybrid2 = True
                session_types[int(cur_day)].hybrid1 = True

        # enumerate the rows so that we can access individual names and last names
        for index,row in enumerate(reader):
            if index == 0 and 'First' in str(row[0]):
                isUsage = False
            elif index == 0 and 'First' not in str(row[0]):
                isUsage = True
    
            # usage report
            if isUsage == True:
                
                # skip the first 2 rows since they just have the column header and my name
                if index == 0 or index == 1:
                    continue
                # Since the first and last names are separated by space, we can split it into 2 different strings. row[0] = first name & row[1] = last name
                row1 = row[0].split()
                # if the cell is blank we don't want to do anything with it
                if len(row1) == 0:
                    continue
                # if they just put a first name just continue
                if len(row1) == 1:
                    continue
                # if more than 2 names total we will only use the first one as the first name and the last one as the last name
                if len(row1) > 2:
                    lenNames = len(row1)
                    # store the first and the last name only
                    row1[1] = str(row1[lenNames - 1])
                # pass the first and last name to the function
                first_name = str(row1[0])
                last_name = str(row1[1])
                cur_day = int(cur_day)
                addToGrid(wb, ws, cur_day, first_name, last_name, isDouble, isReview, isHybrid, True)
            # meeting report
            else:
                
                # skip the first row since it is a header
                if index == 0:
                    continue
                first_name = str(row[0])
                last_name = str(row[1])
                cur_day = int(cur_day)
                addToGrid(wb, ws, cur_day, first_name, last_name, isDouble, isReview, isHybrid, True)  

class NameScreen(QDialog):
    def __init__(self, first_name_1, last_name_1, first_name_2, last_name_2):
        super().__init__()
        self.first_name_1 = first_name_1
        self.last_name_1 = last_name_1
        self.first_name_2 = first_name_2
        self.last_name_2 = last_name_2
        self.initUI()

    def initUI(self):
        self.btn = QPushButton('Show Dialog', self)
        self.btn.move(20,20)
        self.btn.clicked.connect(self.showDialog)

    def showDialog(self):
        text, ok = QInputDialog().getText(self, "Your input is required.", "Do the two names below refer to the same person? y or n\n1) " + self.first_name_1 + ' ' + self.last_name_1 + '\n2) ' + self.first_name_2 + ' ' + self.last_name_2 + '                                                                            ')
        if ok and text:
            global isSame
            isSame = text     

def runAutoGrid():
    # open the xlsx grid file
    wb = load_workbook(grid_name)
    # get the current active worksheet (The only one)
    ws = wb.worksheets[0]
    # get names_not_listed_row, totals_row, and first_blank_row
    global first_blank_row
    global names_not_listed_row
    global totals_row

    for i in range(2, ws.max_row + 1):
        temp = str(ws.cell(row=i, column=1).value)

        if 'not listed' in temp:
            names_not_listed_row = i

        if temp == 'None':
            first_blank_row = i
            break

    totals_row = ws.max_row
    
    # if the user uploaded an in-person attendance sheet
    if inPerson == True:
        # open the xlsx attendance sheet file
        wb2 = load_workbook(attendance_sheet_name)
        ws2 = wb2.worksheets[0]

        last_day = 0
        last_session = '' 
        # store if session number 1 or 2
        last_number = 0
        last_hybrid = False
        # store if the session was hybrid
        # store the names of the sessions in case people fill out the survey out of order
        person_session_1_name = ''
        person_session_1_day = 0
        person_session_2_name = ''
        person_session_2_day = 0
        hybrid_session_1_name = ''
        hybrid_session_1_day = 0
        hybrid_session_2_name = ''
        hybrid_session_2_day = 0
        # need to have it in a loop, since we will be filling in multiple days potentially
        # so first we need to use the attendance sheet file to determine how many days to input for the given month
        # loop through all the entries in the attendance report and count how many we need for this month and keep track of which row this month starts on
        for i in range(2, ws2.max_row + 1):
            temp = str(ws2.cell(row=i, column=2).value)
            # account for if the table has empty slots
            if temp == "None":
                continue
            curMonth = ''
            curDay = ''
            isDouble = 0
            isReview = False
            isHybrid = False
            curYear = ''
            curYear = curYear + temp[0] + temp[1] + temp[2] + temp[3]

            for j in range(0, len(temp)):
                if temp[j] == '-':
                    curMonth = curMonth + temp[j + 1] + temp[j + 2]
                    curDay = curDay + temp[j + 4] + temp[j + 5]
                    break
            # get the integer representing the day of the week using the date (monday = 0)
            weekday = datetime(int(curYear), int(curMonth), int(curDay)).weekday()
            # variable to check if the attendee logged the correct session for that day
            check = True

            # store which session it is so that we can see if there are multiple sessions on that
            # day
            # the stored string is ex. "Monday's Session 10:00 AM - 10:50 AM in CB 119"
            temp = str(ws2.cell(row=i, column=11).value).lower()
            
            # if it is the first session on the table, we will just make last_day and last_session
            # and last_number this session
            if i == 2:
                last_day = int(curDay)
                last_session = temp
                last_number = 1
                if 'tch 226' in temp:
                    last_hybrid = True
                else:
                    last_hybrid = False

            # check if its a review session
            if 'review' in temp:
                isReview = True
                # mark this day's object as having a review in person
                session_types[int(curDay)].review_person = True
            
            # if this is not a review session
            else:
                # check that they have the correct date for the weekday
                if weekdays[weekday] not in temp and 'review' not in temp:
                    check = False
                # check if we are on the same day as the last session with the same session name as
                # the last session
                if last_day == int(curDay) and last_session == temp:
                    # if we are then this session is on the same column as the last session
                    isDouble = last_number
                    # check if the session is hybrid
                    isHybrid = last_hybrid
                # if the day is the same but the sessions are different
                elif last_day == int(curDay) and last_session != temp:
                    # if the attendee chose the wrong session
                    if check == False:
                        # in this case the attendee chose the incorrect session by accident
                        # so we will just revert to whatever the previous session was since it was
                        # the same day.
                        isDouble = last_number
                        isHybrid = last_hybrid
                    # if the attendee chose the correct session
                    else:
                        # this means that we are on a new session on the same day
                        # This can be a few scenarios:
                        # 1) the previous session was not hybrid and now this one is
                        if last_hybrid == False and 'tch 226' in temp:
                            isHybrid = True
                            # if this is the first hybrid session
                            if temp == hybrid_session_1_name and int(curDay) == hybrid_session_1_day:
                                isDouble = 1
                            # if this is the second hybrid session
                            elif temp == hybrid_session_2_name and int(curDay) == hybrid_session_2_day:
                                isDouble = 2
                            # if this is the first hybrid session of the day
                            else:
                                isDouble = 1
                                # store the info of the hybrid session on this day
                                hybrid_session_1_day = int(curDay)
                                hybrid_session_1_name = temp 

                        # 2) the previous session was hybrid and this new session is also hybrid
                        elif last_hybrid == True and 'tch 226' in temp:
                            isHybrid = True
                            # store this as the second hybrid session of the day
                            isDouble = 2
                            hybrid_session_2_day = int(curDay)
                            hybrid_session_2_name = temp 
                        
                        # 3) the previous session was not hybrid and this one is not either
                        elif last_hybrid == False and 'tch 226' not in temp:
                            isHybrid = False
                            # store this as the second in-person session of the day
                            isDouble = 2
                            person_session_2_day = int(curDay)
                            person_session_2_name = temp

                        # 4) the previous session was hybrid and this one is not
                        elif last_hybrid == True and 'tch 226' not in temp:
                            isHybrid = False
                            # if this is the same as first session of the day
                            if temp == person_session_1_name and int(curDay) == person_session_1_day:
                                isDouble = 1
                            # if this is the same as the second session of the day
                            if temp == person_session_2_day and int(curDay) == person_session_2_day:
                                isDouble = 2
                            # if this is the first in-person session of the day
                            else:
                                isDouble = 1
                                # store the info of this first in-person session
                                person_session_1_day = int(curDay)
                                person_session_1_name = temp

                        # store the session info for the next session   
                        last_hybrid = isHybrid
                        last_session = temp
                        last_day = int(curDay)
                        last_number = isDouble

                # if the day is different    
                else:
                    # if the user put in the wrong session we have no way of knowing which session
                    # they belong to
                    if check == False:
                        continue
                    
                    # if the user put in the correct session then just mark it as either the first
                    # person session or the first remote session of the day
                    if 'tch 226' in temp:
                        hybrid_session_1_name = temp
                        hybrid_session_1_day = int(curDay)
                        isHybrid = True
                    else:
                        person_session_1_day = int(curDay)
                        person_session_1_name = temp
                        isHybrid = False
                    isDouble = 1
                    # now store the session info for the next session
                    last_hybrid = isHybrid
                    last_number = isDouble
                    last_session = temp
                    last_day = int(curDay) 
            
            # mark this day's object with applicable scenarios
            # it is a review
            if isReview == True:
                session_types[int(curDay)].review_person = True
            else:
                # it is the first of 2 regular sessions
                if isDouble == 0 and isHybrid == False:
                    session_types[int(curDay)].person1 = True
                # it is the first of 2 regular sessions
                if isDouble == 1 and isHybrid == False:
                    session_types[int(curDay)].person1 = True
                # it is the second of 2 regular sessions
                if isDouble == 2 and isHybrid == False:
                    session_types[int(curDay)].person2 = True
                    session_types[int(curDay)].person1 = True
                # it is the only hybrid session of the day
                if isHybrid == True and isDouble == 0:
                    session_types[int(curDay)].hybrid1 = True
                # it is the first of 2 hybrid sessions
                if isHybrid == True and isDouble == 1:
                    session_types[int(curDay)].hybrid1 = True
                # it is the second of 2 hybrid sessions
                if isHybrid == True and isDouble == 2:
                    session_types[int(curDay)].hybrid2 = True
                    session_types[int(curDay)].hybrid1 = True

            curMonth = int(curMonth)
            curDay = int(curDay)

            if curMonth == current_month:
                first_name = str(ws2.cell(row=i, column=6).value).strip()
                last_name = str(ws2.cell(row=i, column=7).value).strip()
                first = first_name
                last = last_name
                # if more than 2 names total we will only use the first one as the first name and the last one as the last name
                first_name = first_name.split()
                if len(first_name) > 1:
                    first = str(first_name[0])
                last_name = last_name.split()
                if len(last_name) > 1:
                    last = str(last_name[len(last_name) - 1])
                addToGrid(wb, ws, curDay, first, last, isDouble, isReview, isHybrid, False)
    '''
    # every time we add a new column, we make sure all the rows have the correct sum formulas
    updatedColCount = ws.max_column
    for i in range(3, names_not_listed_row):
        colLetter = numToLetter[updatedColCount]
        ws['{}{}'.format(colLetter, i)] = '=SUM({}{}:{}{})'.format(numToLetter[4], i, numToLetter[updatedColCount - 3], i)
    # update the last column that counts how many days had attendance
    updatedColCount = ws.max_column + 1
    ws['{}{}'.format(numToLetter[updatedColCount], names_not_listed_row)] = '=COUNTIF({}{}:{}{},">0")'.format(numToLetter[4], totals_row, numToLetter[updatedColCount-3], totals_row)
    '''
    # save because we are done entering names
    wb.save(grid_name)
    # once we are done filling the colomn colors the program can end


def addExamColor(ws, wb, color, num, days):
    colorFill = PatternFill(start_color=color, fill_type='solid')
    numColumns = ws.max_column - 1
    for i in range(0, num):
        for j in range(4, numColumns):
            # if the column is green then we skip it
            color = ws['{}{}'.format(numToLetter[j], 2)].fill.start_color.index
            if color == 'FF00CC00':
                continue
            # if the column number matches with the day we want to fill
            if str(days[i]) == str(ws.cell(row=2, column=j).value):
                # fill the entire column with the color
                for k in range(2, totals_row):
                    ws.cell(row=k, column=j).fill = colorFill
                wb.save(grid_name)
                # don't break because we want all of them to be red

def addReviewColor(wb, ws, color, col):
    colorFill = PatternFill(start_color=color, fill_type='solid')
    for i in range(2, totals_row):
        ws.cell(row=i, column=col).fill = colorFill
    wb.save(grid_name)

def addToGrid(wb, ws, curDay, first_name, last_name, isDouble, isReview, isHybrid, isZoom):
    # We need to specify that these are global variables before we can edit them in this function
    global first_blank_row
    global totals_row
    global synonym
    global notSynonym
    global session_types
    global used_days

    # mark the day as used if it hasn't already
    if used_days[curDay] == False:
        used_days[curDay] = True

    # Find the current number of columns in the grid
    gridColumnCount = ws.max_column - 1

    # Determine which column the current person needs to go into in the grid based on the current day.
    curCol = -1
    for i in range(4, gridColumnCount):
        # need to make sure that it is still recognized even if it has an r attached to it
        # so instead of comparing them directly, we can just check if the day is in the column
        # title, since we go chronologically
        if str(ws.cell(row=2, column=i).value) == str(curDay):
            curCol = i
            break
    if curCol == -1:
        return
    temp_col = curCol
    # if the 10 columns for this day have not been created yet then create them
    if session_types[curDay].created == False and str(ws.cell(row=2, column=curCol).value) != str(ws.cell(row=2, column=curCol+1).value):
        # mark that day as created
        session_types[curDay].created = True
        # store the first column so we can easily find the correct column
        temp_col = curCol
        # then create them in the following order
        # session 1 P
        ws.cell(row=1, column=curCol).value = 'P'
        
        # use a for loop to create the other 9 columns
        # the dictionary called 'modality' will store the correct modality for each iteration
        for i in range(0, 9):
            # insert a second column
            ws.insert_cols(curCol+1)
            # move into that new column
            curCol += 1
            # add the day at the top of the column, making sure to store it as an integer
            ws.cell(row=2, column=curCol).value = int(ws.cell(row=2, column=curCol-1).value)
            # make the top number bold and have the correct font size of 13
            ws['{}{}'.format(numToLetter[curCol], 2)].font = Font(size = 13, bold=True)
            # label the modality at the top
            ws.cell(row=1, column=curCol).value = modality[i]
            # make the modality bold and have the correct font size of 13
            ws['{}{}'.format(numToLetter[curCol], 1)].font = Font(size = 13, bold=True)
            # add the sum formulas to the 2 rows that store the totals on each column
            # determine the letter of the current column
            colLetter = numToLetter[curCol]
            # sum from row 3 to right before 'names not listed'
            ws['{}{}'.format(colLetter, names_not_listed_row)] = '=SUM({}{}:{}{})'.format(colLetter, 3, colLetter, names_not_listed_row-1)
            # sum from 'names not listed' to the row with the last name on the list
            ws['{}{}'.format(colLetter, totals_row)] = '=SUM({}{}:{}{})'.format(colLetter, names_not_listed_row, colLetter, totals_row -1)
            # give every cell in the column a border
            for i in range(2, totals_row+1):
                # if it is the top row with the number, we want the border to have a thick bottom
                if i == 2:
                    ws.cell(row=i, column=curCol).border = border_thick_bottom
                # otherwise we give it a normal border
                else:
                    ws.cell(row=i, column=curCol).border = border
            # every time we add a new column, we make sure all columns have the correct sum formulas
            updatedColCount = ws.max_column
            for i in range(4, updatedColCount):
                colLetter = numToLetter[i]
                ws['{}{}'.format(colLetter, names_not_listed_row)] = '=SUM({}{}:{}{})'.format(colLetter, 3, colLetter, names_not_listed_row-1)
                ws['{}{}'.format(colLetter, totals_row)] = '=SUM({}{}:{}{})'.format(colLetter, names_not_listed_row, colLetter, totals_row -1)
        # make all of the columns the same width
        for i in range(4, ws.max_column):
            ws.column_dimensions['{}'.format(numToLetter[i])].width = 5
        
    # At this point all modalities for the day are definitely created so determine which one to
    # the current name into
    # Below are the options:
    # in-person & hybrid & isDouble = 0 or 1: HP1
    # in-person & hybrid & isDouble = 2: HP2
    # in-person & not hybrid & isDouble = 0 or 1: P1
    # in-person & not hybrid & isDouble = 2: P2
    # in-person & review: RP

    # zoom & hybrid & isDouble = 0 or 1: HR1
    # zoom & hybrid & isDouble = 2: HR2
    # zoom & not hybrid and isDouble = 0 or 1: R1
    # zoom & not hybrid and isDouble = 2: R2
    # zoom & review: RR
    
    # go to the first column of the day
    curCol = temp_col
    
    if isZoom == False and isReview == True:
        curCol += correct_column['RP']
    elif isZoom == False and isHybrid == True and (isDouble == 0 or isDouble == 1):
        curCol += correct_column['HP1']
    elif isZoom == False and isHybrid == True and isDouble == 2:
        curCol += correct_column['HP2']
    elif isZoom == False and isHybrid == False and (isDouble == 0 or isDouble == 1):
        curCol += correct_column['P1']
    elif isZoom == False and isHybrid == False and isDouble == 2:
        curCol += correct_column['P2']
    
    if isZoom == True and isReview == True:
        curCol += correct_column['RR']
    elif isZoom == True and isHybrid == True and (isDouble == 0 or isDouble == 1):
        curCol += correct_column['HR1']
    elif isZoom == True and isHybrid == True and isDouble == 2:
        curCol += correct_column['HR2']
    elif isZoom == True and isHybrid == False and (isDouble == 0 or isDouble == 1):
        curCol += correct_column['R1']
    elif isZoom == True and isHybrid == False and isDouble == 2:
        curCol += correct_column['R2']
    
    # if the current column is green, it means we alraedy had a review here so we need to move columns
    color = ws['{}{}'.format(numToLetter[curCol], 2)].fill.start_color.index

    # if the column is not green and the person attended a review session
    if color != 'FF00CC00' and isReview == True:
        # make the current column green since it is a review session
        addReviewColor(wb, ws, 'FF00CC00', curCol)

    # if they were observed and the first name is in the mentors dictionary and the value is the same as the last name
    if num_observations != 0 and str(first_name).lower() in mentors and mentors[str(first_name).lower()] == str(last_name).lower():
        # do not put a mentor name in the grid
        return

    # check will change to 1 if the name is found so that we don't add any duplicate names
    check = 0
    reportLast = last_name.lower()
    reportFirst = first_name.lower()
    # remove any symbols from the names of the attendance reports
    for char in string.punctuation:
        reportFirst = reportFirst.replace(char, '')
        reportLast = reportLast.replace(char, '')

    # Compare each name and last name from the attendance file to the grid to add the '1' in the appropriate cell
    for i in range(3, first_blank_row):
        # store the first and last name on each row of the grid into a temp vari
        gridLast = str(ws.cell(row=i, column=1).value).lower()
        gridFirst = str(ws.cell(row=i, column=2).value).lower()
        # remove any symbols from both last names
        for char in string.punctuation:
            gridFirst = gridFirst.replace(char, '')
            gridLast = gridLast.replace(char, '')
        # loop for however many last names there are for that one person on the grid
        tempArr = gridLast.split()
        for j in range(0, len(tempArr)):
            # check if the last name matches, even if the person just put in an initial
            if reportLast == tempArr[j] or (len(reportLast) == 1 and reportLast == tempArr[j][0] and reportFirst[0] == gridFirst[0]) or (len(reportLast) > 1 and reportLast in tempArr[j]) or gridLast.replace(' ', '') == reportLast:
                # check for however many first names there are if they match
                tempFArr = gridFirst.split()
                for k in range(0, len(tempFArr)):
                    if reportFirst == tempFArr[k] or (len(reportFirst) == 1 and reportFirst == tempFArr[k][0]):
                        # add the 1 to the right cell
                        ws.cell(row=i, column=curCol).value = 1
                        # change check to 1 to show that we found the name
                        check = 1
                        # break so we don't keep trying to find a name we already found
                        break
                # if the first name does not match we want to make sure it isn't just a nickname
                if check == 0 and (gridFirst[0] == reportFirst[0] or reportFirst in gridFirst):
                    # check if we have any nicknames for the first name in the grid in our dictionary
                    if str(gridFirst) in synonym: #and synonym[str(reportFirst+reportLast)] == str(gridFirst+tempArr[j]):
                        for k in range(0, len(synonym[gridFirst])):
                            # if we find the nickname from the report as a valid nickname of the grid name
                            if synonym[gridFirst][k] == reportFirst:
                                # add the 1 to the right cell
                                ws.cell(row=i, column=curCol).value = 1
                                # change check to 1 to show that we found the name
                                check = 1
                                # break so we don't keep trying to find a name we already found
                                break
                        # if we don't find the name in the dictionary, we can ask the user if the names are the same
                        if check == 0:
                            #response = input('Is {} {} the same as {} {}? Y or N: '.format(reportFirst, reportLast, gridFirst, tempArr[j]))
                            name_screen = NameScreen(reportFirst, reportLast, gridFirst, tempArr[j])
                            name_screen.showDialog()
                            if str(isSame).lower() == 'yes' or str(isSame).lower() == 'y':
                                # add the 1 to the right cell
                                ws.cell(row=i, column=curCol).value = 1
                                # change check to 1 to show that we found the name
                                check = 1
                                # add the new synonym to the full name key
                                synonym[gridFirst].append(reportFirst)
                                # break so we don't keep trying to find a name we already found
                                break
                            # if the user says that these are not the same people then we add the names to notSynonym
                            else:
                                # add the 2 names so that we don't ask again
                                if gridFirst not in notSynonym:
                                    notSynonym[gridFirst] = list()
                                notSynonym[gridFirst] = reportFirst
                                continue
                    # check if the key is in the notSynonym dictionary and check that the value stored for that key is the same the first name on the grid. if so continue.
                    elif gridFirst in notSynonym: # and notSynonym[str(reportFirst+reportLast)] == str(gridFirst+tempArr[j]):
                        check = 0
                        for k in range(0, len(notSynonym[gridFirst])):
                            # if we find that the name on the report is not a nickname of the name on the grid we can just continue
                            if notSynonym[gridFirst][k] == reportFirst:
                                check = 1
                                break
                        if check == 0:
                            #response = input('Is {} {} the same as {} {}? Y or N: '.format(reportFirst, reportLast, gridFirst, tempArr[j]))
                            name_screen = NameScreen(reportFirst, reportLast, gridFirst, tempArr[j])
                            name_screen.showDialog()
                            if str(isSame).lower() == 'yes' or str(isSame).lower() == 'y':
                                # add the 1 to the right cell
                                ws.cell(row=i, column=curCol).value = 1
                                # change check to 1 to show that we found the name
                                check = 1
                                # add the new synonym to the full name key
                                if gridFirst not in synonym:
                                    synonym[gridFirst] = list()
                                synonym[gridFirst].append(reportFirst)
                                # break so we don't keep trying to find a name we already found
                                break
                            # if the user says that these are not the same people then we add the names to notSynonym
                            else:
                                # add the 2 names so that we don't ask again
                                notSynonym[gridFirst].append(reportFirst)
                                continue
                        if check == 1:
                            continue
                    else:
                        # ask the user to see if the person used a nickname, perhaps. Ex: andy instead of andrew
                        #response = input('Is {} {} the same as {} {}? Y or N: '.format(reportFirst, reportLast, gridFirst, tempArr[j]))
                        name_screen = NameScreen(reportFirst, reportLast, gridFirst, tempArr[j])
                        name_screen.showDialog()
                        if str(isSame).lower() == 'yes' or str(isSame).lower() == 'y':
                            # add the 1 to the right cell
                            ws.cell(row=i, column=curCol).value = 1
                            # change check to 1 to show that we found the name
                            check = 1
                            # add the new value to the dictionary of synonyms
                            synonym[gridFirst] = list()
                            synonym[gridFirst].append(reportFirst)
                            # break so we don't keep trying to find a name we already found
                            break
                        # if the user says that these are not the same people then we add the names to notSynonym
                        else:
                            # add the 2 names so that we don't ask again
                            notSynonym[gridFirst] = list()
                            notSynonym[gridFirst].append(reportFirst)
                            continue

    # if we did not find the name (check = 0) then we add it to the next open row, which was given as a global variable first_blank_row
    if check == 0 and first_blank_row < totals_row:
        print(first_blank_row)
        ws.cell(row=first_blank_row, column=1).value = reportLast # add the last name to the first column
        ws.cell(row=first_blank_row, column=2).value = reportFirst # add the first name to the second column
        ws.cell(row=first_blank_row, column=curCol).value = 1  # add a '1' to the appropriate cell to show attendance from this person
        first_blank_row += 1 # increase the first_blank_row to show that the next open row is now 1 below this new name

    # if we did not find the name but we are on the row with the totals, we don't want to overwrite any of those, so we add a row before first_blank_row before adding the name
    elif check == 0 and first_blank_row == totals_row:
        ws.insert_rows(first_blank_row) # add a new row right before the limit
        ws.cell(row=first_blank_row, column=1).value = reportLast # add the last name
        ws.cell(row=first_blank_row, column=2).value = reportFirst # add the first name
        ws.cell(row=first_blank_row, column=curCol).value = 1 # add a '1' to the appropriate cell to show attendance from this person
        first_blank_row += 1 # increment the last row value by one
        totals_row += 1 # increment the totals row value by 1
        # update the formula of the totals_row
        updatedColCount = ws.max_column
        for i in range(4, updatedColCount):
            colLetter = numToLetter[i]
            ws['{}{}'.format(colLetter, totals_row)] = '=SUM({}{}:{}{})'.format(colLetter, names_not_listed_row, colLetter, totals_row -1)
        # update the last column that counts how many days had attendance
        ws['{}{}'.format(numToLetter[updatedColCount], names_not_listed_row)] ='=COUNTIF({}{}:{}{},">0")'.format(numToLetter[3], totals_row,numToLetter[updatedColCount-2], totals_row)

# main
if __name__ == "__main__":
    # launch the gui
    app = QApplication(sys.argv)
    config.app = app
    app.setWindowIcon(QtGui.QIcon('logo.ico')) # this sets my logo as the window icon

    screen_resolution = app.desktop().screenGeometry()
    width, height = screen_resolution.width(), screen_resolution.height()
    key = str(width) + "x" + str(height)
    startingLocation = []
    if key not in config.res:
        startingLocation = [500, 500]
    else:
        startingLocation = config.res[key]


    mw = FirstWindow.MainWindow()
    mw.show()

    sys.exit(app.exec_())
