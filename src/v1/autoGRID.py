import sys
import os
from PyQt5.uic import loadUi
from PyQt5 import QtWidgets, QtCore, QtGui
from PyQt5.QtWidgets import QDialog, QApplication, QWidget, QFileDialog, QInputDialog, QLineEdit, QPushButton
from openpyxl import load_workbook # for opening xlsx files like the grids
from openpyxl.styles import PatternFill # to fill columns
from openpyxl.styles.borders import Border, Side # to create borders an new columns
from openpyxl.styles import Font # to write the bold number at the top of columns
import csv # for opening csv  files like the attendance reports
import string # for removing symbols from names
import ctypes

# this sets the icon as your taskbar icon
myappid = 'AutoGrid'
ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID(myappid)

# The code below makes it so that the widget doesn't get messed up if you scale windows text by 125% or more
# Query DPI Awareness (Windows 10 and 8)
awareness = ctypes.c_int()
errorCode = ctypes.windll.shcore.GetProcessDpiAwareness(0, ctypes.byref(awareness))
# Set DPI Awareness  (Windows 10 and 8)
errorCode = ctypes.windll.shcore.SetProcessDpiAwareness(0)
# the argument is the awareness level, which can be 0, 1 or 2:
# for 1-to-1 pixel control I seem to need it to be non-zero (I'm using level 2)
# Set DPI Awareness  (Windows 7 and Vista)
success = ctypes.windll.user32.SetProcessDPIAware()
# behaviour on later OSes is undefined, although when I run it on my Windows 10 machine, it seems to work with effects identical to SetProcessDpiAwareness(1)

# create the global variables
current_month = 0
grid_name = ''
attendance_sheet_name = ''
first_blank_row = 0
names_not_listed_row = 0
totals_row = 0
num_observations = 0
num_exams = 0
exam_days = []
mentors = {}
synonym = {}
notSynonym = {}
done_observations = 0
isSame = '' # used to determine nicknames
zoom_attendance_reports = []
num_zoom_attendance_reports = 0
zoom_days = []
inPerson = True # used to determine if they have any inPerson attendance to report

# create dictionary of letters and numbers so we can find the proper letter for a certain column number
numToLetter = {} # numbers are keys and letters are values
# create string of uppercase letters
upperCaseString = string.ascii_uppercase
alphabetList = list(upperCaseString)
for i in range(0, 50):
    # this if statement will give us AA and AB etc up until AX since the grids will never be larger than this
    if i >= 26:
        numToLetter[i+1] = alphabetList[0] + alphabetList[i-26]
    else:
        numToLetter[i+1] = alphabetList[i]

# create border types for when we create new columns
border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
border_thick_bottom = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thick'))

class WelcomeScreen(QDialog):
    def __init__(self):
        super(WelcomeScreen, self).__init__()
        loadUi("welcomescreen.ui", self)
        self.start_button.clicked.connect(self.startProgram)
    # launch the month selector widget when the start button is pressed
    def startProgram(self):
        month = MonthScreen()
        widget.addWidget(month)
        widget.setCurrentIndex(widget.currentIndex()+1)

class ZoomFileScreen(QDialog):
    def __init__(self):
        super(ZoomFileScreen, self).__init__()
        loadUi("zoomfile.ui", self)
        # for browse button
        self.browse_button.clicked.connect(self.browseReports)
        # for next button
        self.next_button_2.clicked.connect(self.nextScreen)

    # function to browse files and return a string list with all of the files
    def browseReports(self):
        global zoom_attendance_reports
        global zoom_days
        # get the names of all the attendance reports
        aTuple = QFileDialog.getOpenFileNames(self, 'open files', '', 'CSV files (*.csv)')
        # place those names in a global list if they selected the correct number of files
        if len(aTuple[0]) == num_zoom_attendance_reports:
            zoom_attendance_reports *= 0
            zoom_days *= 0
            for i in range(0, len(aTuple[0])):
                zoom_attendance_reports.append(aTuple[0][i])
            # now get the day of each file and put it in another list with the same indexes
            for i in range(0, len(zoom_attendance_reports)):
                temp = zoom_attendance_reports[i]
                lastIndex = len(temp) - 1
                # if there is a dash 1 it means it is a double day
                if ('-1' in temp) or ('-2' in temp):
                    # this means it is a double digit date
                    if lastIndex - 7 >= 0 and temp[lastIndex - 7].isdigit():
                        zoom_days.append(temp[lastIndex - 7] + temp[lastIndex - 6] + temp[lastIndex - 5] + temp[lastIndex - 4])
                    # this means it is a single digit date
                    else:
                        zoom_days.append(temp[lastIndex - 6] + temp[lastIndex - 5] + temp[lastIndex - 4])
                # this means it is a review day
                elif temp[lastIndex - 4] == 'r':
                    # we have 2 options
                    # double digit day with an r at the end
                    if (lastIndex - 6) >= 0 and temp[lastIndex - 6].isdigit():
                        zoom_days.append(temp[lastIndex - 6] + temp[lastIndex - 5] + temp[lastIndex - 4])
                    # single digit day with an r at the end
                    else:
                        zoom_days.append(temp[lastIndex - 5] + temp[lastIndex - 4])
                # this means it is a double digit day
                elif (lastIndex - 5) >= 0 and temp[lastIndex - 5].isdigit():
                    zoom_days.append(temp[lastIndex - 5] + temp[lastIndex - 4])
                # this means that it is a single digit date
                else:
                    zoom_days.append(temp[lastIndex - 4])
            self.error_2.setText('All ' + str(num_zoom_attendance_reports) + ' reports received')
    # launch the program widget once they have input all files
    def nextScreen(self):
        errors = 0
        # if they don't select enough files or if they select too many files it doesn't proceed and an error message pops up
        if len(zoom_attendance_reports) != num_zoom_attendance_reports:
            errors += 1
            self.error_2.setText('Please select ' + str(num_zoom_attendance_reports) + ' grid files')
        else:
            self.error_2.setText('')
            program = ProgramScreen()
            widget.addWidget(program)
            widget.setCurrentIndex(widget.currentIndex()+1)

class ProgramScreen(QDialog):
    def __init__(self):
        super(ProgramScreen, self).__init__()
        loadUi("program.ui", self)
        # for the browse buttons
        self.grid_browse.clicked.connect(self.browseFilesGrid)
        self.attendance_sheet_browse.clicked.connect(self.browseFilesAttendanceSheets)
        # when run button is pressed
        self.run_button.clicked.connect(self.runProgram)

    # function for browsing files
    def browseFilesGrid(self):
        fileName=QFileDialog.getOpenFileName(self, 'open file', '','XLSX files (*.xlsx)')
        self.grid_name.setText(fileName[0])
    def browseFilesAttendanceSheets(self):
        ############################################################################################################################################check back on this os.getcwd() vs '/home'
        fileName=QFileDialog.getOpenFileName(self, 'open file', '', 'XLSX files (*.xlsx)')
        self.attendance_sheet_name.setText(fileName[0])
    # function for extracting variables
    def runProgram(self):
        global num_observations
        global num_exams
        global grid_name
        global attendance_sheet_name
        global inPerson
        # need to make sure they fully input all of the data before pressing run so the program doesn't crash
        errors = 0
        # start to take in the input
        # don't proceed if they leave a line blank except for the attendance_sheet_name, which is optional
        if len(self.grid_name.text()) == 0:
            self.grid_error.setText("Please select your grid.")
            errors += 1
        else:
            self.grid_error.setText('')
            grid_name = self.grid_name.text()
        if len(self.attendance_sheet_name.text()) == 0:
            inPerson = False
        else:
            self.attendance_sheet_error.setText('')
            attendance_sheet_name = self.attendance_sheet_name.text()
            inPerson = True
        if len(self.num_observations.text()) == 0:
            self.error_4.setText("Please fill in this field.")
            errors += 1
        else:
            self.error_4.setText('')
            num_observations = int(self.num_observations.text())

        if len(self.num_exams.text()) == 0:
            self.error_6.setText("Please fill in this field.")
            errors += 1
        else:
            self.error_6.setText('')
            num_exams = int(self.num_exams.text())

        # if they have been observed at least once we have to launch the observations page
        global done_observations
        if errors == 0 and num_observations > 0:
            observation_page = ObservationScreen()
            widget.addWidget(observation_page)
            widget.setCurrentIndex(widget.currentIndex()+1)
        elif errors == 0 and int(num_observations) == 0:
            done_observations = 1
        if errors == 0 and num_exams > 0 and done_observations == 1:
            review_page = ReviewScreen()
            widget.addWidget(review_page)
            widget.setCurrentIndex(widget.currentIndex()+1)
        if errors == 0 and num_observations == 0 and num_exams == 0:
            loading_screen = LoadingScreen()
            widget.addWidget(loading_screen)
            widget.setCurrentIndex(widget.currentIndex()+1)

class ObservationScreen(QDialog):
    def __init__(self):
        super(ObservationScreen, self).__init__()
        loadUi("observations.ui", self)
        # once we press the 'Done' button, we want to store all the info of the mentors' names
        self.done_button.clicked.connect(self.storeMentorNames)

    def storeMentorNames(self):
        errors = 0
        global mentors
        if num_observations > 0:
            if len(self.first_name_1.text()) == 0 and len(self.last_name_1.text()) == 0:
                self.error_1a.setText('Please fill in this field')
                self.error_1b.setText('Please fill in this field')
                errors += 2
            elif len(self.first_name_1.text()) == 0 or len(self.last_name_1.text()) == 0:
                if len(self.first_name_1.text()) == 0:
                    self.error_1a.setText('Please fill in this field')
                    self.error_1b.setText('')
                    errors += 1
                if len(self.last_name_1.text()) == 0:
                    self.error_1b.setText('Please fill in this field')
                    self.error_1a.setText('')
                    errors += 1
            else:
                self.error_1a.setText('')
                self.error_1b.setText('')
                mentors[str(self.first_name_1.text())] = str(self.last_name_1.text())
        if num_observations > 1:
            if len(self.first_name_2.text()) == 0 and len(self.last_name_2.text()) == 0:
                self.error_2a.setText('Please fill in this field')
                self.error_2b.setText('Please fill in this field')
                errors += 2
            elif len(self.first_name_2.text()) == 0 or len(self.last_name_2.text()) == 0:
                if len(self.first_name_2.text()) == 0:
                    self.error_2a.setText('Please fill in this field')
                    self.error_2b.setText('')
                    errors += 1
                if len(self.last_name_2.text()) == 0:
                    self.error_2b.setText('Please fill in this field')
                    self.error_2a.setText('')
                    errors += 1
            else:
                self.error_2a.setText('')
                self.error_2b.setText('')
                mentors[str(self.first_name_2.text())] = str(self.last_name_2.text())
        if num_observations > 2:
            if len(self.first_name_3.text()) == 0 and len(self.last_name_3.text()) == 0:
                self.error_3a.setText('Please fill in this field')
                self.error_3b.setText('Please fill in this field')
                errors += 2
            elif len(self.first_name_3.text()) == 0 or len(self.last_name_3.text()) == 0:
                if len(self.first_name_3.text()) == 0:
                    self.error_3a.setText('Please fill in this field')
                    self.error_3b.setText('')
                    errors += 1
                if len(self.last_name_3.text()) == 0:
                    self.error_3b.setText('Please fill in this field')
                    self.error_3a.setText('')
                    errors += 1
            else:
                self.error_3a.setText('')
                self.error_3b.setText('')
                mentors[str(self.first_name_3.text())] = str(self.last_name_3.text())
        if num_observations > 3:
            if len(self.first_name_4.text()) == 0 and len(self.last_name_4.text()) == 0:
                self.error_4a.setText('Please fill in this field')
                self.error_4b.setText('Please fill in this field')
                errors += 2
            elif len(self.first_name_4.text()) == 0 or len(self.last_name_4.text()) == 0:
                if len(self.first_name_4.text()) == 0:
                    self.error_4a.setText('Please fill in this field')
                    self.error_4b.setText('')
                    errors += 1
                if len(self.last_name_4.text()) == 0:
                    self.error_4b.setText('Please fill in this field')
                    self.error_4a.setText('')
                    errors += 1
            else:
                self.error_4a.setText('')
                self.error_4b.setText('')
                mentors[str(self.first_name_4.text())] = str(self.last_name_4.text())
        # check if we need to loop at the review days too
        if errors == 0 and num_exams > 0:
            review_page = ReviewScreen()
            widget.addWidget(review_page)
            widget.setCurrentIndex(widget.currentIndex()+1)
        elif errors == 0 and num_exams == 0:
            loading_screen = LoadingScreen()
            widget.addWidget(loading_screen)
            widget.setCurrentIndex(widget.currentIndex()+1)

class ReviewScreen(QDialog):
    def __init__(self):
        super(ReviewScreen, self).__init__()
        loadUi("reviews.ui", self)
        self.done_button.clicked.connect(self.storeExamInfo)

    def storeExamInfo(self):
        global exam_days
        errors = 0

        # exam days
        if num_exams > 0:
            if len(self.exam_1.text()) == 0:
                self.error_5.setText('Please fill in this field')
                errors += 1
            else:
                self.error_5.setText('')
                exam_days.append(int(self.exam_1.text()))
        if num_exams > 1:
            if len(self.exam_2.text()) == 0:
                self.error_6.setText('Please fill in this field')
                errors += 1
            else:
                self.error_6.setText('')
                exam_days.append(int(self.exam_2.text()))
        if num_exams > 2:
            if len(self.exam_3.text()) == 0:
                self.error_7.setText('Please fill in this field')
                errors += 1
            else:
                self.error_7.setText('')
                exam_days.append(int(self.exam_3.text()))
        if num_exams > 3:
            if len(self.exam_4.text()) == 0:
                self.error_8.setText('Please fill in this field')
                errors += 1
            else:
                self.error_8.setText('')
                exam_days.append(int(self.exam_4.text()))

        # run the grid filling program
        if errors == 0:
            loading_screen = LoadingScreen()
            widget.addWidget(loading_screen)
            widget.setCurrentIndex(widget.currentIndex()+1)

class LoadingScreen(QDialog):
    def __init__(self):
        super(LoadingScreen, self).__init__()
        loadUi("load.ui", self)
        runAutoGrid()
        runAutoGridZoom()
        self.complete()

    def complete(self):
       self.label_2.setText('Your grid has been completed')

class NameScreen(QDialog):
    def __init__(self, first_name_1, last_name_1, first_name_2, last_name_2):
        super().__init__()
        self.first_name_1 = first_name_1
        self.last_name_1 = last_name_1
        self.first_name_2 = first_name_2
        self.last_name_2 = last_name_2
        self.initUI()

    def initUI(self):
        self.btn = QPushButton('Show Dialog', self)
        self.btn.move(20,20)
        self.btn.clicked.connect(self.showDialog)

    def showDialog(self):
        text, ok = QInputDialog().getText(self, "Your input is required.", "Do the two names below refer to the same person? y or n\n1) " + self.first_name_1 + ' ' + self.last_name_1 + '\n2) ' + self.first_name_2 + ' ' + self.last_name_2 + '                                                                            ')
        if ok and text:
            global isSame
            isSame = text

class MonthScreen(QDialog):
    def __init__(self):
        super(MonthScreen, self).__init__()
        loadUi("months.ui", self)
        self.january.clicked.connect(self.jan)
        self.february.clicked.connect(self.feb)
        self.march.clicked.connect(self.mar)
        self.april.clicked.connect(self.apr)
        self.may.clicked.connect(self.mayFunc)
        self.june.clicked.connect(self.juneFunc)
        self.july.clicked.connect(self.julyFunc)
        self.august.clicked.connect(self.aug)
        self.september.clicked.connect(self.sep)
        self.october.clicked.connect(self.oct)
        self.november.clicked.connect(self.nov)
        self.december.clicked.connect(self.dec)

    def jan(self):
        global current_month
        current_month = 1
        zoom = ZoomScreen()
        widget.addWidget(zoom)
        widget.setCurrentIndex(widget.currentIndex()+1)
    def feb(self):
        global current_month
        current_month = 2
        zoom = ZoomScreen()
        widget.addWidget(zoom)
        widget.setCurrentIndex(widget.currentIndex()+1)
    def mar(self):
        global current_month
        current_month = 3
        zoom = ZoomScreen()
        widget.addWidget(zoom)
        widget.setCurrentIndex(widget.currentIndex()+1)
    def apr(self):
        global current_month
        current_month = 4
        zoom = ZoomScreen()
        widget.addWidget(zoom)
        widget.setCurrentIndex(widget.currentIndex()+1)
    def mayFunc(self):
        global current_month
        current_month = 5
        zoom = ZoomScreen()
        widget.addWidget(zoom)
        widget.setCurrentIndex(widget.currentIndex()+1)
    def juneFunc(self):
        global current_month
        current_month = 6
        zoom = ZoomScreen()
        widget.addWidget(zoom)
        widget.setCurrentIndex(widget.currentIndex()+1)
    def julyFunc(self):
        global current_month
        current_month = 7
        zoom = ZoomScreen()
        widget.addWidget(zoom)
        widget.setCurrentIndex(widget.currentIndex()+1)
    def aug(self):
        global current_month
        current_month = 8
        zoom = ZoomScreen()
        widget.addWidget(zoom)
        widget.setCurrentIndex(widget.currentIndex()+1)
    def sep(self):
        global current_month
        current_month = 9
        zoom = ZoomScreen()
        widget.addWidget(zoom)
        widget.setCurrentIndex(widget.currentIndex()+1)
    def oct(self):
        global current_month
        current_month = 10
        zoom = ZoomScreen()
        widget.addWidget(zoom)
        widget.setCurrentIndex(widget.currentIndex()+1)
    def nov(self):
        global current_month
        current_month = 11
        zoom = ZoomScreen()
        widget.addWidget(zoom)
        widget.setCurrentIndex(widget.currentIndex()+1)
    def dec(self):
        global current_month
        current_month = 12
        zoom = ZoomScreen()
        widget.addWidget(zoom)
        widget.setCurrentIndex(widget.currentIndex()+1)

class ZoomScreen(QDialog):
    def __init__(self):
        super(ZoomScreen, self).__init__()
        loadUi("zoom.ui", self)
        self.next_button.clicked.connect(self.getZoomFiles)

    def getZoomFiles(self):
        errors = 0
        global num_zoom_attendance_reports
        if len(self.num_zoom_attendance_sheets.text()) == 0:
            errors += 1
            self.error_1.setText('Please fill in this field')
        else:
            self.error_1.setText('')
            num_zoom_attendance_reports = int(self.num_zoom_attendance_sheets.text())

            if num_zoom_attendance_reports == 0:
                program = ProgramScreen()
                widget.addWidget(program)
                widget.setCurrentIndex(widget.currentIndex()+1)
            else:
                zoomFile = ZoomFileScreen()
                widget.addWidget(zoomFile)
                widget.setCurrentIndex(widget.currentIndex()+1)

def runAutoGridZoom():
    # open the xlsx grid file
    wb = load_workbook(grid_name)
    # get the current active worksheet (The only one)
    ws = wb.worksheets[0]

    # we need to add each name from each csv file to the grid
    for i in range(0, num_zoom_attendance_reports):
        runAutoGridZoomHelper(wb, ws, zoom_attendance_reports[i], zoom_days[i])
    # save because we are done entering names
    wb.save(grid_name)

def runAutoGridZoomHelper(wb, ws, attendance_report, file_day):
    global first_blank_row
    global names_not_listed_row
    global totals_row

    for i in range(2, ws.max_row + 1):
        temp = str(ws.cell(row=i, column=1).value)

        if 'not listed' in temp:
            names_not_listed_row = i

        if temp == 'None':
            first_blank_row = i
            break
    totals_row = ws.max_row

    with open(attendance_report) as fin:
        # read the csv
        reader = csv.reader(fin)
        # get the day out of the name and determine if it is double or not
        isDouble = 0
        # determine if it is a review or not
        isReview = 0
        length = len(file_day)
        if '-1' in file_day:
            isDouble = 1
        if '-2' in file_day:
            isDouble = 2
        if 'r' in file_day:
            isReview = 1
        if 'r' not in file_day and (length == 4 or length == 2):
            cur_day = file_day[0] + file_day[1]
        elif 'r' not in file_day and (length == 3 or length == 1):
            cur_day = file_day[0]
        elif 'r' in file_day and length == 2:
            cur_day = file_day[0]
        elif 'r' in file_day and length == 3:
            cur_day = file_day[0] + file_day[1]
        # enumerate the rows so that we can access individual names and last names
        for index,row in enumerate(reader):
            # skip the first 2 rows since they just have the column header and my name
            if index == 0 or index == 1:
                continue
            # Since the first and last names are separated by space, we can split it into 2 different strings. row[0] = first name & row[1] = last name
            row = row[0].split()
            # if the cell is blank we don't want to do anything with it
            if len(row) == 0:
                continue
            # if they just put a first name just continue
            if len(row) == 1:
                continue
            # if more than 2 names total we will only use the first one as the first name and the last one as the last name
            if len(row) > 2:
                lenNames = len(row)
                # store the first and the last name only
                row[1] = str(row[lenNames - 1])
            # pass the first and last name to the function
            first_name = str(row[0])
            last_name = str(row[1])
            cur_day = int(cur_day)
            addToGrid(wb, ws, cur_day, first_name, last_name, isDouble, isReview)

def runAutoGrid():
    # open the xlsx grid file
    wb = load_workbook(grid_name)
    # get the current active worksheet (The only one)
    ws = wb.worksheets[0]
    # get names_not_listed_row, totals_row, and first_blank_row
    global first_blank_row
    global names_not_listed_row
    global totals_row

    for i in range(2, ws.max_row + 1):
        temp = str(ws.cell(row=i, column=1).value)

        if 'not listed' in temp:
            names_not_listed_row = i

        if temp == 'None':
            first_blank_row = i
            break
    totals_row = ws.max_row


    if inPerson == True:
        # open the xlsx attendance sheet file
        wb2 = load_workbook(attendance_sheet_name)
        ws2 = wb2.worksheets[0]

        # need to have it in a loop, since we will be filling in multiple days potentially
        # so first we need to use the attendance sheet file to determine how many days to input for the given month
        # loop through all the entries in the attendance report and count how many we need for this month and keep track of which row this month starts on
        for i in range(2, ws2.max_row + 1):
            temp = str(ws2.cell(row=i, column=2).value)
            curMonth = ''
            curDay = ''
            isDouble = 0
            isReview = 0
            for j in range(0, len(temp)):
                if temp[j] == '-':
                    curMonth = curMonth + temp[j + 1] + temp[j + 2]
                    curDay = curDay + temp[j + 4] + temp[j + 5]
                    break
            # store which session it is so that we can see if there are multiple sessions on that day
            temp = str(ws2.cell(row=i, column=11).value)
            if '1 of 2' in temp:
                isDouble = 1
            if '2 of 2' in temp:
                isDouble = 2
            if 'Review' in temp:
                isReview = 1
            curMonth = int(curMonth)
            curDay = int(curDay)

            if curMonth == current_month:
                first_name = str(ws2.cell(row=i, column=6).value).strip()
                last_name = str(ws2.cell(row=i, column=7).value).strip()
                first = first_name
                last = last_name
                # if more than 2 names total we will only use the first one as the first name and the last one as the last name
                first_name = first_name.split()
                if len(first_name) > 1:
                    first = str(first_name[0])
                last_name = last_name.split()
                if len(last_name) > 1:
                    last = str(last_name[len(last_name) - 1])
                addToGrid(wb, ws, curDay, first, last, isDouble, isReview)

    # save because we are done entering names
    wb.save(grid_name)
    # once we are done entering the names, we can fill in the columns with colors
    addExamColor(ws, wb, 'FFFF0000', num_exams, exam_days)
    # once we are done filling the colomn colors the program can end

def addExamColor(ws, wb, color, num, days):
    colorFill = PatternFill(start_color=color, fill_type='solid')
    numColumns = ws.max_column - 1
    for i in range(0, num):
        for j in range(3, numColumns):
            # if the column number matches with the day we want to fill
            if str(days[i]) == str(ws.cell(row=2, column=j).value):
                # fill the entire column with the color
                for k in range(2, totals_row):
                    ws.cell(row=k, column=j).fill = colorFill
                wb.save(grid_name)
                # don't break because we want all of them to be red

def addReviewColor(wb, ws, color, col):
    colorFill = PatternFill(start_color=color, fill_type='solid')
    for i in range(2, totals_row):
        ws.cell(row=i, column=col).fill = colorFill
    wb.save(grid_name)

def addToGrid(wb, ws, curDay, first_name, last_name, isDouble, isReview):
    # We need to specify that these are global variables before we can edit them in this function
    global first_blank_row
    global totals_row
    global synonym
    global notSynonym
    # Find the current number of columns in the grid
    gridColumnCount = ws.max_column - 1

    # Determine which column the current person needs to go into in the grid based on the current day.
    curCol = -1
    for i in range(3, gridColumnCount):
        # need to make sure that it is still recognized even if it has an r attached to it
        # so instead of comparing them directly, we can just check if the day is in the column
        # title, since we go chronologically
        if str(ws.cell(row=2, column=i).value) == str(curDay):
            curCol = i
            break
    if curCol == -1:
        return

    # if the current column is green, it means we alraedy had a review here so we need to move columns
    color = ws['{}{}'.format(numToLetter[curCol], 2)].fill.start_color.index
    if color == 'FF00CC00' and isReview == 0:
        # if there is already a second column, we just move into it
        if str(ws.cell(row=2, column=curCol).value) == str(ws.cell(row=2, column=curCol+1).value):
            # just move into it because the isDouble variable will take care of placing us into the correct column
            curCol += 1
        # if there is not a second column, we need to create it and move into it
        elif str(ws.cell(row=2, column=curCol).value) != str(ws.cell(row=2, column=curCol+1).value):
            # insert the new column
            ws.insert_cols(curCol+1)
            # move into that new column
            curCol += 1
            # give every cell in the column a border
            for i in range(2, totals_row):
                # if it is the top row with the number, we want the border to have a thick bottom
                if i == 2:
                    ws.cell(row=i, column=curCol).border = border_thick_bottom
                # otherwise we give it a normal border
                else:
                    ws.cell(row=i, column=curCol).border = border
            # add the day at the top of the column, making sure to store it as an integer
            ws.cell(row=2, column=curCol).value = int(ws.cell(row=2, column=curCol-1).value)
            # make the top number bold and have the correct font size of 13
            ws['{}{}'.format(numToLetter[curCol], 2)].font = Font(size = 13, bold=True)
            # add the sum formulas to the 2 rows that store the totals on each column
            # determine the letter of the current column
            colLetter = numToLetter[curCol]
            # sum from row 3 to right before 'names not listed'
            ws['{}{}'.format(colLetter, names_not_listed_row)] = '=SUM({}{}:{}{})'.format(colLetter, 3, colLetter, names_not_listed_row-1)
            # sum from 'names not listed' to the row with the last name on the list
            ws['{}{}'.format(colLetter, totals_row)] = '=SUM({}{}:{}{})'.format(colLetter, names_not_listed_row, colLetter, totals_row -1)
            # every time we add a new column, we make sure all columns have the correct sum formulas
            updatedColCount = ws.max_column
            for i in range(3, updatedColCount):
                colLetter = numToLetter[i]
                ws['{}{}'.format(colLetter, names_not_listed_row)] = '=SUM({}{}:{}{})'.format(colLetter, 3, colLetter, names_not_listed_row-1)
                ws['{}{}'.format(colLetter, totals_row)] = '=SUM({}{}:{}{})'.format(colLetter, names_not_listed_row, colLetter, totals_row -1)
            # update the last column that counts how many days had attendance
            ws['{}{}'.format(numToLetter[updatedColCount], names_not_listed_row)] = '=COUNTIF({}{}:{}{},">0")'.format(numToLetter[3], totals_row, numToLetter[updatedColCount-2], totals_row)

    # if the column is not green and the person attended a review session
    if color != 'FF00CC00' and isReview == 1:
        isEmpty = True
        # check if this column is empty or not
        for i in range(3, names_not_listed_row - 1):
            if ws.cell(row=i, column=curCol).value != None:
                isEmpty = False
        for i in range(names_not_listed_row + 1, first_blank_row + 1):
            if ws.cell(row=i, column=curCol).value != None:
                isEmpty = False
        # if the column is empty we can place the attendance in this column
        # if the column is not empty, then we need to create a new column behind it, and then move into it
        if isEmpty == False:
            # insert the new column
            ws.insert_cols(curCol)
            # adding a column already moves you into the new column
            # give every cell in the new column a border
            for i in range(2, totals_row):
                # if it is the top row with the number, we want the border to have a thick bottom
                if i == 2:
                    ws.cell(row=i, column=curCol).border = border_thick_bottom
                # otherwise we give it a normal border
                else:
                    ws.cell(row=i, column=curCol).border = border
            # add the day at the top of the column, making sure to store it as an integer
            ws.cell(row=2, column=curCol).value = int(ws.cell(row=2, column=curCol+1).value)
            # make the top number bold and have the correct font size of 13
            ws['{}{}'.format(numToLetter[curCol], 2)].font = Font(size = 13, bold=True)
            # add the sum formulas to the 2 rows that store the totals on each column
            # determine the letter of the current column
            colLetter = numToLetter[curCol]
            # sum from row 3 to right before 'names not listed'
            ws['{}{}'.format(colLetter, names_not_listed_row)] = '=SUM({}{}:{}{})'.format(colLetter, 3, colLetter, names_not_listed_row-1)
            # sum from 'names not listed' to the row with the last name on the list
            ws['{}{}'.format(colLetter, totals_row)] = '=SUM({}{}:{}{})'.format(colLetter, names_not_listed_row, colLetter, totals_row -1)
            # every time we add a new column, we make sure all columns have the correct sum formulas
            updatedColCount = ws.max_column
            for i in range(3, updatedColCount):
                colLetter = numToLetter[i]
                ws['{}{}'.format(colLetter, names_not_listed_row)] = '=SUM({}{}:{}{})'.format(colLetter, 3, colLetter, names_not_listed_row-1)
                ws['{}{}'.format(colLetter, totals_row)] = '=SUM({}{}:{}{})'.format(colLetter, names_not_listed_row, colLetter, totals_row -1)
            # update the last column that counts how many days had attendance
            ws['{}{}'.format(numToLetter[updatedColCount], names_not_listed_row)] = '=COUNTIF({}{}:{}{},">0")'.format(numToLetter[3], totals_row, numToLetter[updatedColCount-2], totals_row)
        # make the current column green since it is a review session
        addReviewColor(wb, ws, 'FF00CC00', curCol)

    # if the color is green and the person attended a review session, we are already in the right place

    # if it is the first session of 2 in the same day
    if isDouble == 1:
        # if there is not a second column, we need to create it even if we are not going to use it yet in case that session had no attendance
        if str(ws.cell(row=2, column=curCol).value) != str(ws.cell(row=2, column=curCol+1).value):
            # insert the new column
            ws.insert_cols(curCol+1)
            # move into that new column
            curCol += 1
            # give every cell in the column a border
            for i in range(2, totals_row):
                # if it is the top row with the number, we want the border to have a thick bottom
                if i == 2:
                    ws.cell(row=i, column=curCol).border = border_thick_bottom
                # otherwise we give it a normal border
                else:
                    ws.cell(row=i, column=curCol).border = border
            # add the day at the top of the column, making sure to store it as an integer
            ws.cell(row=2, column=curCol).value = int(ws.cell(row=2, column=curCol-1).value)
            # make the top number bold and have the correct font size of 13
            ws['{}{}'.format(numToLetter[curCol], 2)].font = Font(size = 13, bold=True)
            # add the sum formulas to the 2 rows that store the totals on each column
            # determine the letter of the current column
            colLetter = numToLetter[curCol]
            # sum from row 3 to right before 'names not listed'
            ws['{}{}'.format(colLetter, names_not_listed_row)] = '=SUM({}{}:{}{})'.format(colLetter, 3, colLetter, names_not_listed_row-1)
            # sum from 'names not listed' to the row with the last name on the list
            ws['{}{}'.format(colLetter, totals_row)] = '=SUM({}{}:{}{})'.format(colLetter, names_not_listed_row, colLetter, totals_row -1)
            # every time we add a new column, we make sure all columns have the correct sum formulas
            updatedColCount = ws.max_column
            for i in range(3, updatedColCount):
                colLetter = numToLetter[i]
                ws['{}{}'.format(colLetter, names_not_listed_row)] = '=SUM({}{}:{}{})'.format(colLetter, 3, colLetter, names_not_listed_row-1)
                ws['{}{}'.format(colLetter, totals_row)] = '=SUM({}{}:{}{})'.format(colLetter, names_not_listed_row, colLetter, totals_row -1)
            # update the last column that counts how many days had attendance
            ws['{}{}'.format(numToLetter[updatedColCount], names_not_listed_row)] = '=COUNTIF({}{}:{}{},">0")'.format(numToLetter[3], totals_row, numToLetter[updatedColCount-2], totals_row)
            # move back to the first column
            curCol -= 1

    # if it is the second session of 2 in the same day
    if isDouble == 2:
        # if it is the second session and there is already a second column, we just move into the second one.
        if str(ws.cell(row=2, column=curCol).value) == str(ws.cell(row=2, column=curCol+1).value):
            curCol += 1
        # if it is the second session of the day and there is only 1 column we go ahead and add the column after the current column
        elif str(ws.cell(row=2, column=curCol).value) != str(ws.cell(row=2, column=curCol+1).value):
            # insert the new column
            ws.insert_cols(curCol+1)
            # move into that new column
            curCol += 1
            # give every cell in the column a border
            for i in range(2, totals_row):
                # if it is the top row with the number, we want the border to have a thick bottom
                if i == 2:
                    ws.cell(row=i, column=curCol).border = border_thick_bottom
                # otherwise we give it a normal border
                else:
                    ws.cell(row=i, column=curCol).border = border
            # add the day at the top of the column, making sure to store it as an integer
            ws.cell(row=2, column=curCol).value = int(ws.cell(row=2, column=curCol-1).value)
            # make the top number bold and have the correct font size of 13
            ws['{}{}'.format(numToLetter[curCol], 2)].font = Font(size = 13, bold=True)
            # add the sum formulas to the 2 rows that store the totals on each column
            # determine the letter of the current column
            colLetter = numToLetter[curCol]
            # sum from row 3 to right before 'names not listed'
            ws['{}{}'.format(colLetter, names_not_listed_row)] = '=SUM({}{}:{}{})'.format(colLetter, 3, colLetter, names_not_listed_row-1)
            # sum from 'names not listed' to the row with the last name on the list
            ws['{}{}'.format(colLetter, totals_row)] = '=SUM({}{}:{}{})'.format(colLetter, names_not_listed_row, colLetter, totals_row -1)
            # every time we add a new column, we make sure all columns have the correct sum formulas
            updatedColCount = ws.max_column
            for i in range(3, updatedColCount):
                colLetter = numToLetter[i]
                ws['{}{}'.format(colLetter, names_not_listed_row)] = '=SUM({}{}:{}{})'.format(colLetter, 3, colLetter, names_not_listed_row-1)
                ws['{}{}'.format(colLetter, totals_row)] = '=SUM({}{}:{}{})'.format(colLetter, names_not_listed_row, colLetter, totals_row -1)
            # update the last column that counts how many days had attendance
            ws['{}{}'.format(numToLetter[updatedColCount], names_not_listed_row)] = '=COUNTIF({}{}:{}{},">0")'.format(numToLetter[3], totals_row, numToLetter[updatedColCount-2], totals_row)

    # make all of the columns the same width
    for i in range(3, ws.max_column):
        ws.column_dimensions['{}'.format(numToLetter[i])].width = 5

    # if they were observed and the first name is in the mentors dictionary and the value is the same as the last name
    if num_observations != 0 and str(first_name).lower() in mentors and mentors[str(first_name).lower()] == str(last_name).lower():
        # do not put a mentor name in the grid
        return

    # check will change to 1 if the name is found so that we don't add any duplicate names
    check = 0
    reportLast = last_name.lower()
    reportFirst = first_name.lower()
    # remove any symbols from the names of the attendance reports
    for char in string.punctuation:
        reportFirst = reportFirst.replace(char, '')
        reportLast = reportLast.replace(char, '')

    # Compare each name and last name from the attendance file to the grid to add the '1' in the appropriate cell
    for i in range(3, first_blank_row):
        # store the first and last name on each row of the grid into a temp vari
        gridLast = str(ws.cell(row=i, column=1).value).lower()
        gridFirst = str(ws.cell(row=i, column=2).value).lower()
        # remove any symbols from both last names
        for char in string.punctuation:
            gridFirst = gridFirst.replace(char, '')
            gridLast = gridLast.replace(char, '')
        # loop for however many last names there are for that one person on the grid
        tempArr = gridLast.split()
        for j in range(0, len(tempArr)):
            # check if the last name matches, even if the person just put in an initial
            if reportLast == tempArr[j] or (len(reportLast) == 1 and reportLast == tempArr[j][0] and reportFirst[0] == gridFirst[0]) or (len(reportLast) > 1 and reportLast in tempArr[j]) or gridLast.replace(' ', '') == reportLast:
                # check for however many first names there are if they match
                tempFArr = gridFirst.split()
                for k in range(0, len(tempFArr)):
                    if reportFirst == tempFArr[k] or (len(reportFirst) == 1 and reportFirst == tempFArr[k][0]):
                        # add the 1 to the right cell
                        ws.cell(row=i, column=curCol).value = 1
                        # change check to 1 to show that we found the name
                        check = 1
                        # break so we don't keep trying to find a name we already found
                        break
                # if the first name does not match we want to make sure it isn't just a nickname
                if check == 0 and (gridFirst[0] == reportFirst[0] or reportFirst in gridFirst):
                    # check if we have any nicknames for the first name in the grid in our dictionary
                    if str(gridFirst) in synonym: #and synonym[str(reportFirst+reportLast)] == str(gridFirst+tempArr[j]):
                        for k in range(0, len(synonym[gridFirst])):
                            # if we find the nickname from the report as a valid nickname of the grid name
                            if synonym[gridFirst][k] == reportFirst:
                                # add the 1 to the right cell
                                ws.cell(row=i, column=curCol).value = 1
                                # change check to 1 to show that we found the name
                                check = 1
                                # break so we don't keep trying to find a name we already found
                                break
                        # if we don't find the name in the dictionary, we can ask the user if the names are the same
                        if check == 0:
                            #response = input('Is {} {} the same as {} {}? Y or N: '.format(reportFirst, reportLast, gridFirst, tempArr[j]))
                            name_screen = NameScreen(reportFirst, reportLast, gridFirst, tempArr[j])
                            name_screen.showDialog()
                            if str(isSame).lower() == 'yes' or str(isSame).lower() == 'y':
                                # add the 1 to the right cell
                                ws.cell(row=i, column=curCol).value = 1
                                # change check to 1 to show that we found the name
                                check = 1
                                # add the new synonym to the full name key
                                synonym[gridFirst].append(reportFirst)
                                # break so we don't keep trying to find a name we already found
                                break
                            # if the user says that these are not the same people then we add the names to notSynonym
                            else:
                                # add the 2 names so that we don't ask again
                                if gridFirst not in notSynonym:
                                    notSynonym[gridFirst] = list()
                                notSynonym[gridFirst] = reportFirst
                                continue
                    # check if the key is in the notSynonym dictionary and check that the value stored for that key is the same the first name on the grid. if so continue.
                    elif gridFirst in notSynonym: # and notSynonym[str(reportFirst+reportLast)] == str(gridFirst+tempArr[j]):
                        for k in range(0, len(notSynonym[gridFirst])):
                            # if we find that the name on the report is not a nickname of the name on the grid we can just continue
                            if notSynonym[gridFirst] == reportFirst:
                                continue
                        #response = input('Is {} {} the same as {} {}? Y or N: '.format(reportFirst, reportLast, gridFirst, tempArr[j]))
                        name_screen = NameScreen(reportFirst, reportLast, gridFirst, tempArr[j])
                        name_screen.showDialog()
                        if str(isSame).lower() == 'yes':
                            # add the 1 to the right cell
                            ws.cell(row=i, column=curCol).value = 1
                            # change check to 1 to show that we found the name
                            check = 1
                            # add the new synonym to the full name key
                            if gridFirst not in synonym:
                                synonym[gridfirst] = list()
                            synonym[gridFirst].append(reportFirst)
                            # break so we don't keep trying to find a name we already found
                            break
                        # if the user says that these are not the same people then we add the names to notSynonym
                        else:
                            # add the 2 names so that we don't ask again
                            notSynonym[gridFirst].append(reportFirst)
                            continue
                    else:
                        # ask the user to see if the person used a nickname, perhaps. Ex: andy instead of andrew
                        #response = input('Is {} {} the same as {} {}? Y or N: '.format(reportFirst, reportLast, gridFirst, tempArr[j]))
                        name_screen = NameScreen(reportFirst, reportLast, gridFirst, tempArr[j])
                        name_screen.showDialog()
                        if str(isSame).lower() == 'yes':
                            # add the 1 to the right cell
                            ws.cell(row=i, column=curCol).value = 1
                            # change check to 1 to show that we found the name
                            check = 1
                            # add the new value to the dictionary of synonyms
                            synonym[gridFirst] = list()
                            synonym[gridFirst].append(reportFirst)
                            # break so we don't keep trying to find a name we already found
                            break
                        # if the user says that these are not the same people then we add the names to notSynonym
                        else:
                            # add the 2 names so that we don't ask again
                            notSynonym[gridFirst] = list()
                            notSynonym[gridFirst].append(reportFirst)
                            continue
    # if we did not find the name (check = 0) then we add it to the next open row, which was given as a global variable first_blank_row
    if check == 0 and first_blank_row < totals_row:
        ws.cell(row=first_blank_row, column=1).value = reportLast # add the last name to the first column
        ws.cell(row=first_blank_row, column=2).value = reportFirst # add the first name to the second column
        ws.cell(row=first_blank_row, column=curCol).value = 1  # add a '1' to the appropriate cell to show attendance from this person
        first_blank_row += 1 # increase the first_blank_row to show that the next open row is now 1 below this new name

    # if we did not find the name but we are on the row with the totals, we don't want to overwrite any of those, so we add a row before first_blank_row before adding the name
    elif check == 0 and first_blank_row == totals_row:
        ws.insert_rows(first_blank_row) # add a new row right before the limit
        ws.cell(row=first_blank_row, column=1).value = reportLast # add the last name
        ws.cell(row=first_blank_row, column=2).value = reportFirst # add the first name
        ws.cell(row=first_blank_row, column=curCol).value = 1 # add a '1' to the appropriate cell to show attendance from this person
        first_blank_row += 1 # increment the last row value by one
        totals_row += 1 # increment the totals row value by 1
        # update the formula of the totals_row
        updatedColCount = ws.max_column
        for i in range(3, updatedColCount):
            colLetter = numToLetter[i]
            ws['{}{}'.format(colLetter, totals_row)] = '=SUM({}{}:{}{})'.format(colLetter, names_not_listed_row, colLetter, totals_row -1)
        # update the last column that counts how many days had attendance
        ws['{}{}'.format(numToLetter[updatedColCount], names_not_listed_row)] = '=COUNTIF({}{}:{}{},">0")'.format(numToLetter[3], totals_row, numToLetter[updatedColCount-2], totals_row)

# main
app = QApplication(sys.argv)
app.setWindowIcon(QtGui.QIcon('logo.ico')) # this sets my logo as the window icon
welcome = WelcomeScreen()
widget = QtWidgets.QStackedWidget()
widget.addWidget(welcome)
widget.setFixedHeight(600)
widget.setFixedWidth(921)
widget.show()
try:
    sys.exit(app.exec())
except:
    print("exiting")
